# -*- coding: utf-8 -*-
"""百度2025年报 财报解读分析表(数据表2) 生成脚本"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

OUTPUT_DIR = r"C:\Users\lishaoming\.qclaw\workspace"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "百度_9888_2025年报_财报解读.xlsx")

# Styles
GREEN_FILL = PatternFill(start_color="006B5A", end_color="006B5A", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="00E8F5F0", end_color="00E8F5F0", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
WHITE_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Microsoft YaHei", size=10, color="333333")
THIN_BORDER = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top', horizontal='left')
CENTER_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='center')

ANALYSIS_DATA = [
    # 一、3分钟看懂版
    [1, "一、3分钟看懂版", "整体判断", "利空(GAAP亏损) + 利好(Non-GAAP盈利+AI高增长)", "数据项1/2/3/11/13/17", "分析"],
    [2, "一、3分钟看懂版", "核心原因1", "GAAP运营亏损58亿: 因核心资产组减值损失162亿(主要是传统搜索业务商誉减值), 说明传统业务价值被重估", "数据项9/13", "分析"],
    [3, "一、3分钟看懂版", "核心原因2", "AI业务爆发式增长: AI业务收入400亿(+34%), Q4占比43%, AI原生营销+301%, 成为新增长引擎", "数据项3", "财报事实"],
    [4, "一、3分钟看懂版", "核心原因3", "毛利率大幅下降: 从51.7%降至43.9%, 说明AI新业务毛利率远低于传统搜索广告, 盈利模式正在转换", "数据项11", "财报事实"],
    # 二、公司靠什么赚钱
    [5, "二、公司靠什么赚钱", "主营业务", "百度核心=在线营销(搜索广告)+非在线营销(AI云+智能驾驶); 爱奇艺=长视频", "数据项3", "财报事实"],
    [6, "二、公司靠什么赚钱", "利润来源", "传统利润来源: 搜索广告(高毛利率60%+)正在萎缩; 新利润来源: AI云服务(毛利率较低但增长快)尚未完全替代", "数据项3/11", "分析"],
    [7, "二、公司靠什么赚钱", "经营变化", "2025年是关键转折: AI收入占一般性业务39%, Q4达43%, AI已成为百度'新核心'; 传统搜索广告持续萎缩; 萝卜快跑海外扩张", "数据项3/4/5", "财报事实"],
    # 三、关键财务数据解读
    [8, "三、关键财务数据解读", "营收", "1291亿元(-3%), 连续2年负增长. 意味着: 传统业务萎缩速度超过AI新业务增长速度, 但AI收入占比快速提升, 转型阵痛期", "数据项17", "分析"],
    [9, "三、关键财务数据解读", "归母净利润", "55.89亿元(GAAP, -76.5%), 受162亿减值拖累. 意味着: 一次性减值影响巨大, 扣除后Non-GAAP净利润189亿元, 实际盈利能力未大幅恶化", "数据项13/17", "分析"],
    [10, "三、关键财务数据解读", "扣非净利润", "Non-GAAP净利润189亿元, 净利率15%. 意味着: 剔除减值等非经常性损益后, 百度仍有稳定盈利能力", "数据项13", "分析"],
    [11, "三、关键财务数据解读", "毛利率", "43.9%(下降7.8个百分点). 意味着: 业务结构正在从高毛利的搜索广告转向较低毛利的AI云/AI应用, 这是战略性转型必然结果", "数据项11", "分析"],
    [12, "三、关键财务数据解读", "净利率", "GAAP净利率4.3%, Non-GAAP净利率15%. 意味着: 真实盈利能力约15%, 但低于互联网行业平均(20%+)", "数据项13/17", "分析"],
    [13, "三、关键财务数据解读", "经营现金流", "-30.13亿元(经营现金流为负). 意味着: 短期现金流承压, 主要因运营资本变动(应收/应付变动)和减值影响", "数据项9", "财报事实"],
    [14, "三、关键财务数据解读", "资产负债率", "35.5%. 意味着: 负债率适中, 财务结构稳健, 有足够空间继续AI投入", "数据项15", "分析"],
    [15, "三、关键财务数据解读", "应收账款", "流动应收账款129.72亿元. 意味着: 占营收约10%, 回款情况尚可, 但销售及管理费用中预期信用损失增加需关注", "数据项15", "分析"],
    [16, "三、关键财务数据解读", "存货", "367.83亿元(主要为爱奇艺内容版权). 意味着: 存货占比高, 爱奇艺内容成本仍是负担", "数据项15", "分析"],
    [17, "三、关键财务数据解读", "ROE", "GAAP ROE仅2.1%, Non-GAAP ROE约7.1%. 意味着: 无论是GAAP还是Non-GAAP, 资本回报率都偏低, 说明公司处于投入期", "数据项13", "分析"],
    [18, "三、关键财务数据解读", "分红/回购/融资", "首次采纳股息政策(预计2026年首次派发); 50亿美元回购计划; 无大规模融资. 意味着: 公司从'只投入不分红'转向'回馈股东', 是成熟期信号", "数据项21", "财报事实"],
    # 四、识别隐藏风险
    [19, "四、识别隐藏风险", "利润质量", "GAAP净利润55.89亿 vs Non-GAAP净利润189亿, 差异133亿(主要是162亿减值). 意味着: 减值是否充分? 如果传统业务继续恶化, 可能还有进一步减值空间", "数据项13/17", "分析"],
    [20, "四、识别隐藏风险", "现金流恶化", "经营现金流-30.13亿元(首次转负), 下半年虽转正(39亿)但全年仍为负. 意味着: 现金流恶化需要密切关注, 如果持续为负将影响AI投入能力", "数据项9", "分析"],
    [21, "四、识别隐藏风险", "资产风险", "1. 核心资产组减值162亿(商誉减值); 2. 爱奇艺存货367.83亿(内容版权); 3. 长期投资449.18亿(公允价值波动风险); 4. 递延税项资产45.82亿(需未来利润抵扣). 意味着: 资产质量有隐患", "数据项15", "分析"],
    [22, "四、识别隐藏风险", "一次性收益", "2025年没有明显的一次性收益美化业绩, 反而是162亿一次性减值损失严重拖累利润. 但需注意: 利息收入86.02亿元(占税前利润的128%), 说明利润对利息收入依赖度高", "数据项9", "分析"],
    # 五、未来展望
    [23, "五、未来展望", "增长逻辑", "1. AI云收入持续高增长(+34%, Q4高性能计算+143%); 2. 文心大模型日均调用量持续爆发; 3. 萝卜快跑全球扩张; 4. 昆仑芯分拆上市(估值重估机会)", "数据项3/5/9", "分析"],
    [24, "五、未来展望", "风险因素", "1. 传统搜索广告持续萎缩(份额被抖音/微信/小红书侵蚀); 2. AI云竞争激烈(阿里云/华为云); 3. 毛利率持续下降; 4. 经营现金流为负; 5. AI投入回报不确定", "数据项4/8/11", "分析"],
    [25, "五、未来展望", "跟踪指标", "1. AI业务收入占比(目标: 超过50%); 2. 经营现金流是否转正; 3. 毛利率是否企稳; 4. 萝卜快跑商业化进展; 5. 昆仑芯分拆上市进展; 6. 文心大模型调用量增速", "数据项3/9/11", "推测"],
    # 六、投资者视角结论
    [26, "六、投资者视角结论", "长期投资者", "AI转型是百度的生死之战, 目前AI收入占比快速提升(39%→43%)是积极信号. 但毛利率下降+现金流为负是隐忧. PB仅0.85x(破净), 如果AI转型成功, 长期空间较大; 如果失败, 当前估值也不贵", "数据项2/11/13/19", "分析"],
    [27, "六、投资者视角结论", "短期投资者", "短期催化剂: 昆仑芯分拆上市+首次股息+50亿美元回购. 短期风险: 2026Q1/Q2业绩可能继续承压(传统业务萎缩+AI投入). 目标价349.30美元(38家机构均值), 当前252.64美元有38%上行空间", "数据项2/5/21", "分析"],
    [28, "六、投资者视角结论", "机会 vs 风险", "更像是'高风险高回报'的机会: AI业务高增长是机会, 但传统业务萎缩+盈利模式转换+现金流为负是风险. 关键观察点: AI收入占比能否在2026年超过50%, 经营现金流能否转正", "数据项3/9/11/17", "分析"],
]

def create_analysis_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "百度2025财报解读分析"
    
    # Column widths
    ws.column_dimensions['A'].width = 5.0
    ws.column_dimensions['B'].width = 18.0
    ws.column_dimensions['C'].width = 18.0
    ws.column_dimensions['D'].width = 65.0
    ws.column_dimensions['E'].width = 20.0
    ws.column_dimensions['F'].width = 12.0
    
    # Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "百度集团(BIDU/9888) 2025年年报 财报解读分析表(6维度)"
    cell.font = TITLE_FONT
    cell.fill = GREEN_FILL
    cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 45
    
    # Headers
    headers = ["序号", "维度", "分析项", "内容", "依据", "类型"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = WHITE_FONT
        cell.fill = GREEN_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 30
    
    # Data
    for idx, row_data in enumerate(ANALYSIS_DATA):
        row = idx + 3
        ws.row_dimensions[row].height = 80
        
        for col, val in enumerate(row_data):
            cell = ws.cell(row=row, column=col+1, value=val)
            cell.font = DATA_FONT
            cell.alignment = WRAP_ALIGN if col >= 3 else CENTER_ALIGN
            cell.border = THIN_BORDER
            cell.fill = WHITE_FILL
    
    ws.freeze_panes = 'A3'
    wb.save(OUTPUT_FILE)
    print(f"[OK] Analysis Excel saved to: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_analysis_excel()
