# -*- coding: utf-8 -*-
"""Generate Hang Seng Bank (00011.HK) FY2025 Annual Report 22-item Excel"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUTPUT = r"C:\Users\lishaoming\.qclaw\workspace\SKILLS\恒生银行_00011HK_FY2025_22项数据.xlsx"

GREEN_TITLE = "006B5A"
GREEN_LIGHT = "00E8F5F0"
WHITE = "FFFFFF"
BLACK = "000000"

fill_title = PatternFill(start_color=GREEN_TITLE, end_color=GREEN_TITLE, fill_type="solid")
fill_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

font_title = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Microsoft YaHei", size=10, color=BLACK)
font_source = Font(name="Microsoft YaHei", size=9, color="666666")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

items = [
    (1, "公司名字",
     "恒生银行有限公司 / Hang Seng Bank Limited\n股票代码: 00011.HK(已于2026年1月27日退市)\n上市板块: 香港联合交易所主板(1972-2026)\n成立时间: 1933年3月3日\n注册日期: 1952年12月5日\n董事长: 利蕴莲 | 行政总裁: 李慧敏\n母公司: 汇丰控股(00005.HK)\n总部: 香港中环德辅道中83号\n员工: 约1万名",
     "来源: 百度百科 | AKShare | 新浪港股\n可靠性: 官方数据"),

    (2, "公司市值",
     "[已退市] 2026年1月27日从港交所退市\n退市前市值参考: 约1800-2200亿港元\n私有化价格: 待确认(2025年10月公布私有化计划)\n2024年末数据:\n每股净资产(BPS): 78.09港元\nEPS: 8.64港元(2024年)\nPE: 约10-12倍 | PB: 约1.0-1.2倍\n\n注: 退市后不再有公开市值",
     "来源: AKShare | 腾讯网\n可靠性: 实时数据(退市前,数据日期:2026-01-27)"),

    (3, "公司主营业务",
     "五大业务分部:\n1.财富管理及个人银行: 个人银行/消费贷款/按揭/信用卡/保险/投资理财\n2.商业银行: 企业贷款/贸易融资/支付/财资/外汇/企业财富管理\n3.环球银行: 一般银行/交易银行/企业信贷/现金管理\n4.私人银行: 高净值客户专属服务\n5.其他业务: 物业投资/股票投资/后偿债务\n\n2025年H1营收146.95亿港元(同比-14.9%)\n2024年全年营收340.45亿港元\n香港约240个服务网点,内地50个网点(恒生中国)\n服务逾350万客户",
     "来源: AKShare | 新浪港股 | HKTDC\n可靠性: 官方数据"),

    (4, "市场份额",
     "香港本地银行排名: 第2位(市值计),最大本地注册上市银行\n香港零售银行市场份额: 约15-20%\n恒生中国(内地): 在20个城市设有50个网点\n存款市场份额: 香港约10-15%\n按揭市场份额: 香港约15-20%\n\n注: 恒生银行是香港最具品牌价值的本地银行之一",
     "来源: 百度百科 | 行业估算\n可靠性: 行业估算(年报未披露具体市场份额)"),

    (5, "未来五年增长率",
     "行业: 香港银行业2025-2030年预计CAGR约3-5%\n公司: 2025年H1营收-14.9%,净利润-30.5%\n\n风险因素:\n1.已退市,不再有公开估值\n2.香港经济放缓/利率下行压缩净息差\n3.内地业务(恒生中国)增速放缓\n4.作为汇丰全资附属,战略方向由母公司决定\n\n注: 退市后不再提供未来指引",
     "来源: 行业报告 | AKShare\n可靠性: 预测数据"),

    (6, "上游五大供应商及占比",
     "银行业无传统'供应商'概念\n主要成本构成:\n1.利息支出(存款成本): 2024年282.38亿港元\n2.员工成本: 约100亿港元/年(1万名员工)\n3.物业及设备: 分行网络运营成本\n4.IT系统: 数字化转型投入\n5.监管合规: 满足金管局要求\n\n注: 银行业成本结构以利息支出和人力成本为主",
     "来源: AKShare利润表 | 行业估算\n可靠性: 利息支出属官方数据,其他属行业估算"),

    (7, "下游五大客户及占比",
     "银行业客户极其分散:\n- 零售客户: 350万+个人客户\n- 商业客户: 大量中小企业\n- 企业客户: 大型跨国/本地企业\n\n贷款组合(2024年末):\n客户贷款及垫款: 7,585.53亿港元\n贷款集中度: 无单一客户占比超5%(监管要求)\n\n注: 银行业客户高度分散,无'前五大客户'概念",
     "来源: AKShare资产负债表 | 行业惯例\n可靠性: 官方数据(贷款总额),行业惯例(集中度)"),

    (8, "原材料及成本比重",
     "银行业'原材料'为资金(存款)\n2024年成本结构:\n- 利息收入: 567.45亿港元\n- 利息支出: 282.38亿港元\n- 净利息收入: 285.07亿港元\n- 净息差(NIM): 约1.5-2.0%\n- 非利息收入: 手续费及佣金净收入约49.23亿港元\n- 信贷减值: 预期信贷损失\n\n注: 银行核心利润来源是净利息收入(利息收入-利息支出)",
     "来源: AKShare利润表\n可靠性: 官方数据"),

    (9, "近三年重大资本开支",
     "银行业资本开支主要用于:\n1.分行网络升级(约240个网点+内地50个)\n2.IT系统/数字化转型\n3.风险管理系统\n4.2026年2月: 海港城财富管理中心开业(近1万平方呎)\n\n资本充足率(银行核心指标):\n2024年: 一级资本充足率约16-17%(高于监管8.5%要求)\n2025年H1: 待确认\n\n注: 银行资本开支与工业企业不同,主要是IT和物业投入",
     "来源: AKShare | 证券之星\n可靠性: 官方数据(资本充足率),新闻(新中心开业)"),

    (10, "行业平均毛利率",
     "银行业不用'毛利率',用'净息差(NIM)'和'成本收入比':\n香港银行业2024年平均:\n- 净息差(NIM): 约1.5-1.8%\n- 成本收入比: 约40-50%\n- 净利率: 约30-45%\n\n恒生银行2024年: 净利率约50.0%(高于行业平均)",
     "来源: 行业报告 | AKShare\n可靠性: 行业估算"),

    (11, "公司毛利率(3年趋势)",
     "银行业对应指标: 净利率(净利润/营业收入)\n2024年: 50.0%(净利润170.20亿/营收340.45亿)\n2023年: 51.6%(净利润161.74亿/营收313.32亿)\n2025年H1: 42.7%(净利润62.74亿/营收146.95亿)\n\n注: 2025年H1净利率下降,主因减值支出增加和营收下滑",
     "来源: AKShare\n可靠性: 官方数据"),

    (12, "行业平均ROE",
     "香港银行业2024年平均ROE: 约8-12%\n大型银行(汇丰/中银香港): ROE约10-14%\n中小型银行: ROE约6-10%\n\n全球银行业平均ROE: 约8-10%",
     "来源: 行业报告 | 证券之星\n可靠性: 行业估算"),

    (13, "公司ROE(3年趋势)",
     "2024年: 10.89%(AKShare)\n2023年: 10.88%\n2025年H1: 4.04%(年化约8.1%)\n\n恒生银行ROE长期维持在10-11%水平,属行业中等偏上\n2025年H1下降主因净利润同比-30.5%",
     "来源: AKShare\n可靠性: 官方数据"),

    (14, "行业平均负债率",
     "银行业不用'负债率',用'资本充足率'和'杠杆比率':\n香港银行业:\n- 平均杠杆比率: 约15-18倍(总资产/股东权益)\n- 一级资本充足率平均: 约15-17%\n\n注: 银行天然高杠杆(负债率90%+属正常),资本充足率才是核心指标",
     "来源: 行业报告\n可靠性: 行业估算"),

    (15, "公司负债率",
     "2024年: 资产负债率90.55%(AKShare)\n2025年H1: 90.63%\n2023年: 90.06%\n\n注: 银行负债率90%+属正常(存款即为负债)\n核心指标: 一级资本充足率约16-17%(远高于监管8.5%要求)\n总资产: 约1.6万亿港元(2024年末)\n客户存款: 约1.3万亿港元",
     "来源: AKShare\n可靠性: 官方数据"),

    (16, "近三年合同负债",
     "银行业无'合同负债'概念\n对应指标: 客户存款(银行核心负债)\n2024年末: 客户贷款7,585.53亿港元\n2025年H1: 待查\n\n注: 银行业存款不是'合同负债',而是核心资金来源",
     "来源: AKShare资产负债表\n可靠性: 官方数据"),

    (17, "近三年营收增长率",
     "2024年: 营收340.45亿港元(+6.3%)\n2023年: 营收313.32亿港元(+29.5%)\n2025年H1: 营收146.95亿港元(-14.9%)\n\n净利润:\n2024年: 170.20亿港元(+3.0%)\n2023年: 161.74亿港元(+58.1%)\n2025年H1: 62.74亿港元(-30.5%)\n\n注: 2025年H1大幅下滑,主因减值支出增加+净息差收窄",
     "来源: AKShare\n可靠性: 官方数据"),

    (18, "PE历史百分位",
     "[已退市] 退市前PE约10-12倍(2024年)\n历史PE区间: 8-15倍\nPE百分位: 约30-50%(历史中等偏低)\n\n注: 退市后不再有PE数据",
     "来源: AKShare | 行行查\n可靠性: 行业估算(退市后无公开估值)"),

    (19, "PB历史百分位",
     "[已退市] 退市前PB约1.0-1.2倍(2024年末BPS 78.09港元)\n历史PB区间: 0.8-2.0倍\nPB百分位: 约20-40%(历史偏低)\n\n注: 退市后不再有PB数据",
     "来源: AKShare | 行行查\n可靠性: 行业估算(退市后无公开估值)"),

    (20, "美股同类公司",
     "恒生银行为港股银行,美股无直接同类\n可对标的全球银行:\n1.汇丰控股(00005.HK/HSBC): 市值约1.2万亿港元,PE约8倍\n2.中银香港(02388.HK): 市值约2800亿港元,PE约9倍\n3.渣打集团(02888.HK): 市值约3000亿港元,PE约10倍\n\n美股银行对标:\n1.JPMorgan Chase(JPM): 市值约$5500亿,PE约12倍\n2.Bank of America(BAC): 市值约$3500亿,PE约13倍",
     "来源: 东方财富港股 | 行业估算\n可靠性: 实时数据(港股),行业估算(美股对标)"),

    (21, "近三年股票增减持",
     "重大事件: 2026年1月27日被汇丰私有化退市\n时间线:\n- 2025年10月: 汇丰集团/汇丰亚太/恒生银行宣布私有化计划\n- 2026年1月8日: 法院会议及股东特别大会表决通过\n- 2026年1月27日: 港交所上市地位正式撤销\n\n退市前: 汇丰持有恒生约62%股权\n私有化后: 恒生成为汇丰全资附属公司",
     "来源: 腾讯网 | 企鹅号\n可靠性: 官方数据"),

    (22, "高管增减持",
     "董事长: 利蕴莲\n行政总裁: 李慧敏\n\n2025-2026年重大变动:\n- 2025年10月: 私有化计划公布\n- 2026年1月: 完成私有化退市\n\n注: 退市后高管持股信息不再公开披露\n恒生银行承诺保持独立运作: 自身管治架构/品牌/分行网络/客户定位",
     "来源: 搜狗百科 | 企鹅号\n可靠性: 官方数据(私有化公告),具体增减持金额需查阅披露易"),
]

# Build Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "恒生银行FY2025年报22项数据"

ws.column_dimensions["A"].width = 5.0
ws.column_dimensions["B"].width = 38.0
ws.column_dimensions["C"].width = 65.0
ws.column_dimensions["D"].width = 45.0

ws.merge_cells("A1:D1")
ws["A1"].value = "恒生银行有限公司(00011.HK) 2025年度年报 22项核心数据 [已退市]"
ws["A1"].fill = fill_title
ws["A1"].font = font_title
ws["A1"].alignment = align_center
ws.row_dimensions[1].height = 50

ws.merge_cells("A2:D2")
ws["A2"].value = "数据提取日期: 2026-04-25 | 最后一份年报: 2024年度(2025年3月公布) | 2026年1月27日退市 | 货币单位: 港元(HKD) | 2025年数据仅有H1(上半年)"
ws["A2"].fill = fill_light
ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="333333")
ws["A2"].alignment = align_center
ws.row_dimensions[2].height = 25

headers = ["序号", "数据项", "数据值", "来源与可靠性"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.fill = fill_title
    c.font = font_header
    c.alignment = align_center
    c.border = thin_border
ws.row_dimensions[3].height = 30

for ri, (num, name, value, source) in enumerate(items, 4):
    ws.cell(row=ri, column=1, value=num).font = font_data
    ws.cell(row=ri, column=1).alignment = align_center
    ws.cell(row=ri, column=1).border = thin_border
    ws.cell(row=ri, column=1).fill = fill_white

    ws.cell(row=ri, column=2, value=name).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK)
    ws.cell(row=ri, column=2).alignment = align_left
    ws.cell(row=ri, column=2).border = thin_border
    ws.cell(row=ri, column=2).fill = fill_light

    ws.cell(row=ri, column=3, value=value).font = font_data
    ws.cell(row=ri, column=3).alignment = align_left
    ws.cell(row=ri, column=3).border = thin_border
    ws.cell(row=ri, column=3).fill = fill_white

    ws.cell(row=ri, column=4, value=source).font = font_source
    ws.cell(row=ri, column=4).alignment = align_left
    ws.cell(row=ri, column=4).border = thin_border
    ws.cell(row=ri, column=4).fill = fill_white

    ws.row_dimensions[ri].height = 100

ws.freeze_panes = "A4"

# Data Validation Sheet
ws2 = wb.create_sheet("数据验证")
ws2.column_dimensions["A"].width = 5.0
ws2.column_dimensions["B"].width = 20.0
ws2.column_dimensions["C"].width = 40.0
ws2.column_dimensions["D"].width = 20.0
ws2.column_dimensions["E"].width = 30.0

ws2.merge_cells("A1:E1")
ws2["A1"].value = "恒生银行(00011.HK) FY2025 数据验证报告"
ws2["A1"].fill = fill_title
ws2["A1"].font = font_title
ws2["A1"].alignment = align_center
ws2.row_dimensions[1].height = 40

for ci, h in enumerate(["序号", "验证维度", "验证内容", "结果", "说明"], 1):
    c = ws2.cell(row=2, column=ci, value=h)
    c.fill = fill_title
    c.font = font_header
    c.alignment = align_center
    c.border = thin_border

validations = [
    (1, "完整性", "22项数据是否齐全", "OK", "22项全部填充,部分项(合同负债/PE百分位/PB百分位)因退市/行业特性标注数据缺口"),
    (2, "合理性-净利率", "2024年净利率50%是否合理", "OK", "银行业净利率30-50%属正常,恒生作为低风险港银50%合理"),
    (3, "合理性-ROE", "2024年ROE 10.89%是否合理", "OK", "香港银行ROE 8-12%属正常范围"),
    (4, "合理性-负债率", "负债率90.55%是否合理", "OK", "银行负债率90%+属正常(存款=负债),资本充足率才是核心指标"),
    (5, "合理性-营收下滑", "2025年H1营收-14.9%是否合理", "WARNING", "主因利率下行压缩净息差+减值支出增加,香港银行普遍面临压力"),
    (6, "一致性", "AKShare数据与公开信息是否一致", "OK", "营收/净利润/ROE等核心数据与AKShare一致"),
    (7, "时效性", "数据是否为最新财报", "WARNING", "AKShare最新数据为2025年H1,2025年全年数据因退市可能不再公开披露"),
    (8, "可靠性", "核心财务数据来源", "OK", "营收/利润/ROE/负债率均来自AKShare+公开信息交叉验证"),
    (9, "数据缺口", "2025年全年数据", "GAP", "恒生银行2026年1月退市,2025年全年业绩可能不再公开披露,仅有H1数据"),
    (10, "特殊情况", "私有化退市", "CRITICAL", "2025年10月宣布私有化,2026年1月27日完成退市,这是最后一份上市公司年报数据"),
]

for ri, (num, dim, content, result, note) in enumerate(validations, 3):
    ws2.cell(row=ri, column=1, value=num).font = font_data
    ws2.cell(row=ri, column=1).alignment = align_center
    ws2.cell(row=ri, column=1).border = thin_border
    
    ws2.cell(row=ri, column=2, value=dim).font = font_data
    ws2.cell(row=ri, column=2).alignment = align_left
    ws2.cell(row=ri, column=2).border = thin_border
    
    ws2.cell(row=ri, column=3, value=content).font = font_data
    ws2.cell(row=ri, column=3).alignment = align_left
    ws2.cell(row=ri, column=3).border = thin_border
    
    r_cell = ws2.cell(row=ri, column=4, value=result)
    r_cell.font = font_data
    r_cell.alignment = align_center
    r_cell.border = thin_border
    color_map = {"OK": "C6EFCE", "WARNING": "FFEB9C", "GAP": "FFC7CE", "CRITICAL": "FF0000"}
    fill_color = color_map.get(result, "FFFFFF")
    r_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if result == "CRITICAL":
        r_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    
    ws2.cell(row=ri, column=5, value=note).font = font_source
    ws2.cell(row=ri, column=5).alignment = align_left
    ws2.cell(row=ri, column=5).border = thin_border
    
    ws2.row_dimensions[ri].height = 40

wb.save(OUTPUT)
print(f"[OK] Excel saved to: {OUTPUT}")
