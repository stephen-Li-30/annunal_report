# -*- coding: utf-8 -*-
"""annual-report-extractor smoke test

离线验证核心生成链路：
1. 生成数据表1
2. 生成数据表2
3. 运行验证器并内嵌数据验证Sheet
4. 检查表1 22项、表2 27项、验证Sheet存在
"""
import os
import subprocess
import sys
import tempfile

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from generate_excel_template import create_empty_template, generate_annual_report_excel
from generate_analysis_report import generate_analysis_report


def assert_true(condition, message):
    if not condition:
        raise AssertionError(message)


def main():
    with tempfile.TemporaryDirectory(prefix='annual_report_smoke_') as tmp_dir:
        table1 = os.path.join(tmp_dir, 'smoke_22项数据.xlsx')
        table2 = os.path.join(tmp_dir, 'smoke_财报解读.xlsx')
        report = os.path.join(tmp_dir, 'smoke_22项数据_验证报告.xlsx')

        data = create_empty_template()
        fixtures = {
            1: ('测试公司（TEST）；年报发布状态：annual_report_released', '来源：测试官方年报 | 可靠性：官方数据'),
            2: ('总市值：1000亿元；PE：20；PB：3', '来源：AKShare实时行情接口 | 可靠性：实时数据'),
            3: ('主营业务：测试业务', '来源：AKShare公开财务接口 | 可靠性：官方/权威数据'),
            11: ('30.00%', '来源：AKShare公开财务接口 | 可靠性：官方/权威数据'),
            13: ('12.00%', '来源：AKShare公开财务接口 | 可靠性：官方/权威数据'),
            15: ('40.00%', '来源：AKShare公开财务接口 | 可靠性：官方/权威数据'),
            17: ('营收规模：100.00亿元；营收增长率：10.00%', '来源：AKShare公开财务接口 | 可靠性：官方/权威数据'),
        }
        for seq, (value, source) in fixtures.items():
            data[seq - 1][2] = value
            data[seq - 1][3] = source

        generate_annual_report_excel(
            company_name_cn='烟测公司',
            company_name_en='Smoke Test Corp',
            stock_code='TEST',
            fiscal_year=2025,
            currency_unit='亿元人民币',
            data_list=data,
            output_path=table1,
            data_date='2026-04-24',
        )
        generate_analysis_report(
            company_name_cn='烟测公司',
            company_name_en='Smoke Test Corp',
            stock_code='TEST',
            fiscal_year=2025,
            data_items=data,
            output_path=table2,
            data_date='2026-04-24',
        )

        validator = os.path.join(SCRIPT_DIR, 'data_validator.py')
        result = subprocess.run(
            [sys.executable, validator, table1],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=120,
        )
        assert_true(result.returncode == 0, result.stderr or result.stdout)
        assert_true(os.path.exists(table1), '数据表1未生成')
        assert_true(os.path.exists(table2), '数据表2未生成')
        assert_true(os.path.exists(report), '验证报告未生成')

        wb1 = openpyxl.load_workbook(table1, data_only=True)
        assert_true('数据验证' in wb1.sheetnames, '数据表1缺少数据验证Sheet')
        ws1 = wb1[wb1.sheetnames[0]]
        data_rows = [row for row in ws1.iter_rows(min_row=4, max_row=25, values_only=True) if row[0] is not None]
        assert_true(len(data_rows) == 22, f'数据表1应为22项, 实际{len(data_rows)}项')
        wb1.close()

        wb2 = openpyxl.load_workbook(table2, data_only=True)
        ws2 = wb2.active
        analysis_rows = [row for row in ws2.iter_rows(min_row=4, values_only=True) if row[0] is not None]
        assert_true(len(analysis_rows) == 27, f'数据表2应为27项, 实际{len(analysis_rows)}项')
        wb2.close()

        print('[OK] smoke test passed')
        print(f'  table1: {table1}')
        print(f'  table2: {table2}')
        print(f'  report: {report}')


if __name__ == '__main__':
    main()
