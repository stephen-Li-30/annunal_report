# -*- coding: utf-8 -*-
"""
中航直升机股份有限公司 2025年报 22项数据生成脚本
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 22项年报数据结构
data = [
    {
        "数据项编号": 1,
        "数据项名称": "公司名字",
        "数据值": "中航直升机股份有限公司（中直股份）",
        "数据来源/说明": "公司官网 | 证券之星 | 可靠性：官方数据"
    },
    {
        "数据项编号": 2,
        "数据项名称": "公司市值",
        "数据值": "总市值：274亿元（2026-03）；流通市值：223亿元；PE(动)：约42倍；PB：1.58倍",
        "数据来源/说明": "东方财富 | 搜狐证券 | 可靠性：实时数据（数据日期：2026-03-27）"
    },
    {
        "数据项编号": 3,
        "数据项名称": "公司主营业务",
        "数据值": "航空产品及零部件的开发、设计、研制、生产、销售；民用直升机整机（AC系列）；通用飞机（运12系列）；航空转包生产及客户化服务",
        "数据来源/说明": "搜狐证券 | 同花顺 | 可靠性：官方数据"
    },
    {
        "数据项编号": 4,
        "数据项名称": "市场份额",
        "数据值": "国内直升机制造业主力军，市场地位突出；国内唯一具备直升机和定翼机研制能力的航空制造企业",
        "数据来源/说明": "公司官网 | 行业估算（国内直升机行业集中度较高，无公开市场份额数据）"
    },
    {
        "数据项编号": 5,
        "数据项名称": "未来增长率",
        "数据值": "受益低空经济发展；军机升级换代需求；民机市场拓展；新能源飞行器布局",
        "数据来源/说明": "公司公告 | 行业分析（预测数据，需跟踪实际订单）"
    },
    {
        "数据项编号": 6,
        "数据项名称": "上游供应商",
        "数据值": "航空原材料、发动机、系统件等供应商（具体名单年报未披露）",
        "数据来源/说明": "年报未披露供应商具体名单 | 可靠性：行业估算"
    },
    {
        "数据项编号": 7,
        "数据项名称": "下游客户",
        "数据值": "军方、政府客户、企业客户；国内主要直升机运营商",
        "数据来源/说明": "年报未披露客户具体名单及占比 | 可靠性：行业估算"
    },
    {
        "数据项编号": 8,
        "数据项名称": "原材料成本",
        "数据值": "航空产品成本占收入约90.67%；原材料成本占比高",
        "数据来源/说明": "搜狐财经年报解读 | 可靠性：官方数据"
    },
    {
        "数据项编号": 9,
        "数据项名称": "资本开支",
        "数据值": "2024年完成重大资产重组，募集配套资金约30亿元；收购昌飞集团100%股权、哈飞集团100%股权",
        "数据来源/说明": "公司公告 | 可靠性：官方数据"
    },
    {
        "数据项编号": 10,
        "数据项名称": "行业平均毛利率",
        "数据值": "航空装备行业毛利率约10-15%",
        " 可靠性：行业估算"
    },
    {
        "数据项编号": 11,
        "数据项名称": "公司毛利率",
        "数据值": "9.22%（2025年报）；去年同期10.47%",
        "数据来源/说明": "搜狐证券年报解读 | 可靠性：官方数据"
    },
    {
        "数据项编号": 12,
        "数据项名称": "行业平均ROE",
        "数据值": "航空装备行业ROE约2-5%",
        "数据来源/说明": "行业估算"
    },
    {
        "数据项编号": 13,
        "数据项名称": "公司ROE",
        "数据值": "3.99%（2025年年报）；去年同期约3.7%",
        "数据来源/说明": "证券之星 | 可靠性：官方数据"
    },
    {
        "数据项编号": 14,
        "数据项名称": "行业平均负债率",
        "数据值": "军工行业负债率约60-70%",
        "数据来源/说明": "行业估算"
    },
    {
        "数据项编号": 15,
        "数据项名称": "公司负债率",
        "数据值": "约65%（2025年Q1）；去年同期68.62%",
        "数据来源/说明": "新浪财经 | 可靠性：官方数据"
    },
    {
        "数据项编号": 16,
        "数据项名称": "近三年合同负债",
        "数据值": "（合同负债数据未单独披露，预收款项约XX亿元）",
        "数据来源/说明": "年报未单独披露 | 可靠性：数据缺口"
    },
    {
        "数据项编号": 17,
        "数据项名称": "营收增长率",
        "数据值": "2025年：-2.28%（营收290.86亿元）；2024年：+11.93%；2023年：+9.6%",
        "数据来源/说明": "搜狐证券 | 格隆汇 | 可靠性：官方数据"
    },
    {
        "数据项编号": 18,
        "数据项名称": "PE历史百分位",
        "数据值": "约42倍（动PE），历史PE区间20-100倍",
        "数据来源/说明": "雪球 | 同花顺 | 可靠性：实时数据"
    },
    {
        "数据项编号": 19,
        "数据项名称": "PB历史百分位",
        "数据值": "1.58倍，历史PB区间1.0-3.0倍",
        "数据来源/说明": "中财网 | 可靠性：实时数据"
    },
    {
        "数据项编号": 20,
        "数据项名称": "美股同类公司",
        "数据值": "洛克希德·马丁（NYSE:LMT）、波音（NYSE:BA）、诺斯罗普·格鲁曼（NOC）、德事隆（TEX）",
        "数据来源/说明": "行业对标 | 可靠性：行业估算"
    },
    {
        "数据项编号": 21,
        "数据项名称": "股票增减持",
        "数据值": "2025年Q1：十大流通股东4新进4退出3增持1减持；控股股东中航科工增持计划5000万-1亿元",
        "数据来源/说明": "新浪财经 | 公司公告 | 可靠性：官方数据"
    },
    {
        "数据项编号": 22,
        "数据项名称": "高管增减持",
        "数据值": "近年无高管增持/减持公告记录",
        "数据来源/说明": "万隆证券网 | 可靠性：官方数据"
    },
]

# 创建DataFrame
df = pd.DataFrame(data)

# 创建Excel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "中直股份 FY2025 22项数据"

# 样式
TITLE_FILL = PatternFill("solid", fgColor="006B5A")
HEADER_FILL = PatternFill("solid", fgColor="00E8F5F0")
DATA_FILL = PatternFill("solid", fgColor="FFFFFF")
TITLE_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="000000")
DATA_FONT = Font(name="Arial", size=10, color="000000")

# 标题行
ws.merge_cell("A1:D1")
ws["A1"] = "中航直升机股份有限公司（600038.SH）FY2025 年报22项核心数据"
ws["A1"].fill = TITLE_FILL
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# 日期行
ws.merge_cell("A2:D2")
ws["A2"] = "数据日期：2026-04-25 | 数据来源：搜狐/新浪/东方财富/公司公告 | 财年：FY2025（截止2025年12月31日）"
ws["A2"].font = Font(name="Arial", size=9, color="666666")

# 表头
headers = ["序号", "数据项名称", "数据值", "数据来源/说明"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

# 数据
for row_idx, row_data in enumerate(data, 4):
    ws.cell(row=row_idx, column=1, value=row_data["数据项编号"]).font = DATA_FONT
    ws.cell(row=row_idx, column=2, value=row_data["数据项名称"]).font = DATA_FONT
    ws.cell(row=row_idx, column=3, value=row_data["数据值"]).font = DATA_FONT
    ws.cell(row=row_idx, column=4, value=row_data["数据来源/说明"]).font = DATA_FONT

# 列宽
ws.column_dimensions["A"].width = 5.0
ws.column_dimensions["B"].width = 38.0
ws.column_dimensions["C"].width = 65.0
ws.column_dimensions["D"].width = 45.0

# 行高
for row in range(4, 26):
    ws.row_dimensions[row].height = 100

# 冻结
ws.freeze_panes = "A4"

# 保存
output_path = r"C:\Users\lishaoming\.qclaw\workspace\中直股份_FY2025_22项数据.xlsx"
wb.save(output_path)
print(f"OK: {output_path}")