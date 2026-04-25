# -*- coding: utf-8 -*-
"""Update CQ Beer 600132 Excel with AKShare precise data"""
import os
from openpyxl import load_workbook

output_dir = os.path.join(os.path.dirname(__file__), "..", "examples", "outputs")
f1 = os.path.join(output_dir, "重庆啤酒_600132_2025年报_22项数据.xlsx")

wb = load_workbook(f1)
ws = wb["22项年报数据"]

# Key updates from AKShare data
updates = {
    # Row 4+idx = row number for each item (1-indexed after headers)
    # Row 2 (item 2): 市值 - update ROE and other calculated values
    6: (3, "总股本: 4.84亿股(全流通)\n流通股本: 4.84亿股\n总市值: 约262.31亿元(2026-04-14)\nPE(TTM): 21.10倍; 扣非PE: 21.87倍\nPB: 9.08倍\n最新股价: 约53.66元(2026-04-14)\n每股收益: 2.54元(2025年度)"),
    # Row 7 (item 4): 市场份额 - update with more precise data
    9: (3, "中国啤酒行业第四大啤酒公司(按销量)\n前四大:华润雪花(约30%)、青岛啤酒(约22%)、百威亚太(约18%)、燕京啤酒(约8%)\n重庆啤酒(嘉士伯中国)市占率约7-8%\n嘉士伯集团为全球三大啤酒公司之一\n在重庆/新疆/宁夏/云南等地区市占率领先"),
    # Row 11 (item 8): 原材料成本
    13: (3, "营业成本114.58亿元(2025年),占营收77.8%\n主要原材料:包装物约50%,原材料(麦芽/大米/啤酒花)约25%,制造费用约15%,人工约10%\n2025年毛利率50.88%,成本控制良好"),
    13: (4, "来源级别:第1级+第5级 | 来源:公司年报毛利率(AKShare) + 行业通用成本结构推算 | 可靠性:官方数据(毛利率)+行业估算(成本结构)"),
    # Row 13 (item 10): 行业平均毛利率
    15: (3, "啤酒行业(A股上市公司)平均毛利率约35-45%\n重庆啤酒毛利率50.88%(2025年),显著高于行业平均\n青岛啤酒毛利率约38%,华润啤酒约40%,百威亚太约52%"),
    # Row 14 (item 11): 公司毛利率 - PRECISE from AKShare
    14: (3, "2025年:综合毛利率50.88%(AKShare官方数据)\n2024年:48.57%\n2023年:49.15%\n毛利率持续提升,高端化战略成效显著\n分品牌:国际品牌53.82%(+0.52pp),本土品牌50.92%(+3.37pp)"),
    14: (4, "来源级别:第1级 | 来源:公司2025年年度报告 + AKShare(stock_financial_abstract) | 可靠性:官方/监管披露数据"),
    # Row 16 (item 12): 行业平均ROE
    16: (3, "啤酒行业(A股上市公司)平均ROE约12-18%\n重庆啤酒ROE 81.68%(2025年),远超行业平均\n极高ROE因嘉士伯品牌授权轻资产模式+高净利率+高资产周转率"),
    # Row 17 (item 13): 公司ROE - PRECISE from AKShare
    17: (3, "2025年ROE = 81.68%(AKShare官方数据,年报口径)\n2024年ROE = 65.95%\n2023年ROE = 67.05%\n2022年ROE = 99.69%\nROE极高且波动大,主要因轻资产模式(嘉士伯品牌授权)+高净利率+高资产周转率\n归母净资产仅28.60亿(2025年末),资产基数极低导致ROE极高"),
    17: (4, "来源级别:第1级 | 来源:AKShare(stock_financial_abstract)官方年报数据 | 可靠性:官方/监管披露数据"),
    # Row 18 (item 14): 行业平均负债率
    18: (3, "啤酒行业(A股上市公司)平均资产负债率约45-55%\n重庆啤酒资产负债率73.24%(2025年),显著高于行业平均\n高负债率因嘉士伯品牌授权模式:公司将品牌相关资产和负债集中,净资产基数极低"),
    # Row 19 (item 15): 公司负债率 - PRECISE from AKShare
    19: (3, "2025年资产负债率: 73.24%(AKShare官方数据)\n2024年: 77.63%\n2023年: 70.52%\n负债率持续下降,但仍处于行业较高水平\n高负债率不代表财务风险:主要因嘉士伯品牌授权轻资产模式,经营现金流充沛(26.24亿元)"),
    19: (4, "来源级别:第1级 | 来源:AKShare(stock_financial_abstract)官方年报数据 | 可靠性:官方/监管披露数据"),
    # Row 20 (item 16): 合同负债
    20: (3, "AKShare未直接披露合同负债数据\n啤酒行业合同负债主要为经销商预付款\n公司经营现金流26.24亿元(2025年),现金流充沛\n合同负债数据需查阅年报资产负债表明细"),
    # Row 21 (item 17): 营收增长率 - PRECISE
    21: (3, "2025年: 营收147.22亿元(+0.53%); 归母净利润12.31亿元(+10.43%); 扣非净利润11.88亿元(-2.78%)\n2024年: 营收146.45亿元(-1.14%); 归母净利润11.15亿元(-16.61%); 扣非净利润12.22亿元(-7.00%\n2023年: 营收148.15亿元(+5.53%); 归母净利润13.37亿元(+0.53%); 扣非净利润13.14亿元(-7.01%\n经营现金流: 2025年26.24亿, 2024年25.42亿, 2023年30.97亿\n分红:2025年全年分红12.1亿元(含税),分红率98.30%"),
    21: (4, "来源级别:第1级 | 来源:公司2025年年度报告 + AKShare(stock_financial_abstract) | 可靠性:官方/监管披露数据"),
}

for row_num, (col, new_value) in updates.items():
    ws.cell(row=row_num, column=col).value = new_value

# Also update subtitle row with AKShare note
ws.cell(row=2, column=1).value = "数据来源: 公司2025年年度报告(2026-03-10发布) + AKShare + ProSearch全网搜索 | 货币单位: 人民币 | 数据日期: 2026-04-25"

# Update validation sheet
ws_val = wb["数据验证"]
ws_val.cell(row=11, column=2).value = "数据缺口项:6/7/9/16/18/19(共6项),需后续全网搜索修复"

wb.save(f1)
print(f"[OK] Updated with AKShare data: {f1}")
