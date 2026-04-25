# -*- coding: utf-8 -*-
"""Generate Hang Seng Bank (00011.HK) FY2025 Financial Analysis Report (Table 2)"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUTPUT = r"C:\Users\lishaoming\.qclaw\workspace\SKILLS\恒生银行_00011HK_FY2025_财报解读.xlsx"

GREEN_TITLE = "006B5A"; GREEN_LIGHT = "00E8F5F0"; WHITE = "FFFFFF"; BLACK = "000000"
fill_title = PatternFill(start_color=GREEN_TITLE, end_color=GREEN_TITLE, fill_type="solid")
fill_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
font_title = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Microsoft YaHei", size=10, color=BLACK)
font_source = Font(name="Microsoft YaHei", size=9, color="666666")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"), top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

analysis = [
    (1, "一、3分钟看懂版", "整体判断",
     "重大事件驱动。恒生银行(00011.HK)已于2026年1月27日被汇丰集团私有化退市，这是其最后一份上市公司年报/半年报。2024年业绩稳健(净利润170.2亿港元，ROE 10.89%)，但2025年H1明显承压(营收-14.9%，净利润-30.5%)。退市后作为汇丰全资附属继续运营，品牌独立。",
     "数据项2/13/17/21", "分析"),

    (2, "一、3分钟看懂版", "核心原因",
     "1)2025年H1业绩下滑主因：美联储降息周期压缩净息差(NIM)\n2)减值支出增加(经济前景不确定性)\n3)非利息收入下降(资本市场波动)\n4)汇丰私有化：看中恒生的香港本地品牌+客户基础+资产价值\n5)退市价格反映恒生价值被低估(PB约1.0-1.2倍)",
     "数据项3/11/17", "分析"),

    (3, "二、公司靠什么赚钱", "主营业务",
     "恒生银行是香港最大的本地注册银行(1933年创立)：\n1.财富管理及个人银行：零售银行/按揭/信用卡/保险/理财\n2.商业银行：企业贷款/贸易融资/支付/外汇\n3.环球银行：交易银行/企业信贷/现金管理\n4.私人银行：高净值客户\n5.其他：物业投资/股票投资\n\n香港约240个网点，内地50个网点(恒生中国)，服务350万+客户。",
     "数据项3", "财报事实"),

    (4, "二、公司靠什么赚钱", "利润来源",
     "银行核心利润=净利息收入+非利息收入-减值支出-运营成本\n2024年：\n- 净利息收入：285.07亿港元(利息收入567.45亿-利息支出282.38亿)\n- 非利息收入：手续费及佣金净收入约49.23亿港元\n- 净利率：50.0%\n\n净利息收入是绝对核心(占比约85%)。",
     "数据项3/8", "财报事实"),

    (5, "二、公司靠什么赚钱", "经营变化",
     "2025年最大变化：\n1.被汇丰私有化退市(1月27日完成)\n2.2025年H1营收-14.9%，净利润-30.5%\n3.净息差收窄(利率下行)\n4.2026年2月海港城财富管理中心开业(扩大财富管理)\n5.2026年4月推出'存折宝'(数字化创新)\n\n退市后加速布局财富管理+数字化转型。",
     "数据项3/9/17", "财报事实"),

    (6, "三、关键财务数据解读", "营收",
     "2024年营收340.45亿港元(+6.3%)，2025年H1营收146.95亿港元(-14.9%)。\n\n营收下滑主因：\n1.美联储2024-2025年持续降息，银行净息差(NIM)收窄\n2.贷款需求放缓(香港经济增速放缓)\n3.非利息收入下降(资本市场波动影响财富管理收入)\n\n银行业绩与利率周期高度相关。",
     "数据项17", "分析"),

    (7, "三、关键财务数据解读", "归母净利润",
     "2024年净利润170.20亿港元(+3.0%)\n2025年H1净利润62.74亿港元(-30.5%)\n\n净利润下滑幅度(30.5%)远大于营收(-14.9%)，说明：\n1.减值支出增加(经济不确定性上升)\n2.运营杠杆(收入下降但固定成本不降)\n3.利率下行对利息收入的边际冲击更大",
     "数据项13/17", "分析"),

    (8, "三、关键财务数据解读", "扣非净利润",
     "港股银行通常不区分扣非净利润。恒生银行利润主要来自核心银行业务(净利息收入+手续费)，非经常性损益占比较低。\n\n需关注：预期信贷损失(ECL)变动对利润的影响，经济下行期ECL增加可能显著侵蚀利润。",
     "数据项17", "推测"),

    (9, "三、关键财务数据解读", "净利率(替代毛利率)",
     "银行业用净利率替代毛利率：\n2024年：50.0%(净利润170.2亿/营收340.45亿)\n2023年：51.6%\n2025年H1：42.7%\n\n50%净利率在银行业属优秀水平(汇丰约25%，中银香港约40%)。2025年H1下降7个百分点需关注。",
     "数据项11", "分析"),

    (10, "三、关键财务数据解读", "成本收入比",
     "银行业核心效率指标：\n恒生银行成本收入比约40-45%(行业优秀水平)\n\n2025年H1成本收入比可能上升(收入下降但成本刚性)，需关注全年数据。",
     "数据项10/11", "分析"),

    (11, "三、关键财务数据解读", "经营现金流",
     "2024年每股经营现金流78.05港元(强劲)\n2025年H1每股经营现金流-8.25港元(转负)\n\n银行现金流波动大是常态(存款/贷款规模变化导致)。单期经营现金流为负不一定是问题，但需持续关注。",
     "数据项17", "分析"),

    (12, "三、关键财务数据解读", "资本充足率(替代负债率)",
     "银行核心指标：\n2024年一级资本充足率约16-17%(远高于监管8.5%要求)\n2024年资产负债率90.55%(银行正常水平)\n\n资本充足率充足，意味着银行有能力承受一定程度的资产损失。恒生作为审慎经营的港银，资本缓冲充裕。",
     "数据项15", "财报事实"),

    (13, "三、关键财务数据解读", "贷款质量",
     "银行业核心风险指标：\n- 不良贷款率(NPL)：需查阅年报(恒生通常<2%)\n- 预期信贷损失(ECL)：2025年H1增加(经济下行预期)\n- 贷款组合：按揭占比大(香港房价波动影响)\n\n贷款质量是银行估值的关键变量。",
     "数据项3/7", "推测"),

    (14, "三、关键财务数据解读", "存款基础",
     "恒生银行是香港存款市场份额最高的本地银行之一：\n- 客户存款约1.3万亿港元(2024年末)\n- 存款基础稳定(逾半港人使用)\n- 低成本存款(活期/储蓄)占比高=净息差优势\n\n存款是银行最核心的'原材料'，恒生的存款基础是其最大竞争优势。",
     "数据项3/16", "财报事实"),

    (15, "三、关键财务数据解读", "ROE",
     "2024年ROE 10.89%，2025年H1仅4.04%(年化约8.1%)\n\nROE下降主因净利润下滑。10-11%ROE在银行业属中等偏上，但2025年可能回落至8-9%。\n\n退市后ROE不再公开，但汇丰年报中可能披露恒生分部数据。",
     "数据项13", "分析"),

    (16, "三、关键财务数据解读", "分红/回购/融资",
     "恒生银行以高派息著称(派息率约60-70%)：\n- 2024年每股派息约5.5-6.0港元\n- 股息率约4-5%(退市前)\n\n退市后：作为汇丰全资附属，不再有公开分红。利润上缴汇丰集团。\n\n汇丰私有化恒生的一个可能动机：获取恒生稳定现金流(年利润170亿港元)。",
     "数据项2/17", "分析"),

    (17, "四、识别隐藏风险", "利润质量",
     "2025年H1利润质量下降：\n1.营收-14.9%但净利润-30.5%，运营杠杆放大了收入下滑的冲击\n2.净利率从50%降至42.7%，盈利效率恶化\n3.减值支出增加(经济前景不确定)\n4.经营现金流为负(-8.25港元/股)\n\n但：恒生作为审慎经营的港银，利润'注水'风险低(银行监管严格)。",
     "数据项11/13/17", "分析"),

    (18, "四、识别隐藏风险", "利率风险",
     "银行最大风险之一：利率波动\n1.美联储降息=净息差收窄=利息收入减少\n2.香港联系汇率制度=香港利率跟随美国\n3.2025年预计继续降息，恒生净利息收入可能进一步下降\n4.但：降息也可能减少减值支出(经济改善+借款人还款压力减轻)\n\n利率风险是双向的，但短期内对恒生偏负面。",
     "数据项5/17", "分析"),

    (19, "四、识别隐藏风险", "资产风险",
     "银行主要资产风险：\n1.按揭贷款(房价下跌风险)：香港房价2023-2025年持续调整\n2.商业地产贷款(空置率上升)\n3.内地贷款(恒生中国)：经济放缓+房地产风险\n4.投资组合(债券/股票公允价值波动)\n\n恒生按揭贷款抵押充足(LTV偏低)，但需关注商业地产风险。",
     "数据项3/7", "分析"),

    (20, "四、识别隐藏风险", "私有化估值",
     "汇丰私有化恒生的估值：\n- 退市前PB约1.0-1.2倍\n- 2024年末BPS 78.09港元\n- 私有化价格待确认(可能接近BPS或小幅溢价)\n\n汇丰为何私有化？\n1.恒生PB长期偏低(1.0-1.2倍)，市场低估其价值\n2.私有化后汇丰可100%享有恒生现金流(年利润170亿港元)\n3.消除关联交易/合规成本\n4.恒生品牌/客户/资产价值>市值",
     "数据项2/18/19", "分析"),

    (21, "四、识别隐藏风险", "易忽略点",
     "1.恒生退市后，汇丰年报中将披露恒生分部数据——这是追踪恒生的唯一公开渠道\n2.2025年H1净利率降至42.7%，如果全年维持，ROE可能降至8%以下\n3.恒生中国(内地业务)风险敞口需关注(房地产+地方政府融资)\n4.海港城财富管理中心开业=聚焦高净值客户(战略转型)\n5.'存折宝'=照顾老年客户(差异化竞争)\n6.退市后恒生不再发布独立年报，信息透明度下降",
     "数据项3/9/17", "分析"),

    (22, "五、未来展望", "增长逻辑",
     "退市后增长逻辑改变：\n1.不再追求独立市值最大化，而是为汇丰贡献利润\n2.财富管理是增长引擎(香港富裕人群+大湾区)\n3.数字化转型(存折宝/手机银行)\n4.恒生中国可借助汇丰内地网络扩张\n5.降息周期结束后净息差有望回升",
     "数据项3/5", "分析"),

    (23, "五、未来展望", "风险因素",
     "1.利率持续下行=净息差持续收窄\n2.香港经济放缓=贷款需求+资产质量压力\n3.内地房地产风险(恒生中国敞口)\n4.退市后信息透明度下降(小股东无法跟踪)\n5.汇丰可能整合恒生业务(品牌独立性存疑)\n6.金融科技竞争(虚拟银行/支付宝等)",
     "数据项5/17", "分析"),

    (24, "五、未来展望", "跟踪指标",
     "退市后如何跟踪恒生：\n1.汇丰控股(00005.HK)年报中的恒生分部数据\n2.恒生银行官网(可能继续发布业绩公告)\n3.金管局数据(香港银行体系统计)\n4.关注：净息差/不良贷款率/成本收入比/资本充足率",
     "数据项3/17", "分析"),

    (25, "六、投资者视角结论", "原有股东",
     "恒生原有股东已被私有化(现金退出)。退市价格是否合理？\n- PB约1.0-1.2倍，接近BPS 78.09港元\n- 考虑恒生年利润170亿港元+ROE 10.89%，PB 1.0-1.2倍偏低\n- 如果独立上市，合理PB应为1.3-1.5倍(对应101-117港元/股)\n\n汇丰以较低价格私有化恒生，原有股东可能未获得充分价值。",
     "数据项2/13/18", "分析"),

    (26, "六、投资者视角结论", "汇丰股东",
     "对汇丰(00005.HK)股东而言，私有化恒生是利好：\n1.100%享有恒生年利润170亿港元(此前仅62%)\n2.消除关联交易成本+合规成本\n3.恒生品牌+客户+资产价值被低估，私有化=捡便宜\n4.恒生现金流可支持汇丰派息\n\n建议汇丰股东关注：汇丰年报中恒生分部的业绩贡献。",
     "数据项2/21", "分析"),

    (27, "六、投资者视角结论", "机会 vs 风险",
     "恒生已退市，投资者无法直接投资。但可通过以下方式间接参与：\n1.买入汇丰控股(00005.HK)——间接持有恒生\n2.关注港股同类银行：中银香港(02388.HK)、渣打(02888.HK)\n\n对汇丰股东：私有化恒生是确定性利好(价值低估资产纳入100%)\n对原恒生股东：私有化价格可能偏低，已无选择权\n\n总体：恒生是优质港银资产，汇丰以合理偏低价格私有化，汇丰股东受益。",
     "数据项2/20", "分析"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "恒生银行FY2025财报解读"
ws.column_dimensions["A"].width = 5.0; ws.column_dimensions["B"].width = 18.0; ws.column_dimensions["C"].width = 18.0; ws.column_dimensions["D"].width = 65.0; ws.column_dimensions["E"].width = 18.0; ws.column_dimensions["F"].width = 12.0

ws.merge_cells("A1:F1")
ws["A1"].value = "恒生银行有限公司(00011.HK) 2025年度财报解读分析表 [已退市]"
ws["A1"].fill = fill_title; ws["A1"].font = font_title; ws["A1"].alignment = align_center; ws.row_dimensions[1].height = 50

ws.merge_cells("A2:F2")
ws["A2"].value = "基于2024年全年+2025年H1数据 | 分析日期: 2026-04-25 | 2026年1月27日退市(汇丰私有化)"
ws["A2"].fill = fill_light; ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="333333"); ws["A2"].alignment = align_center; ws.row_dimensions[2].height = 25

for ci, h in enumerate(["序号", "维度", "分析项", "内容", "依据", "类型"], 1):
    c = ws.cell(row=3, column=ci, value=h); c.fill = fill_title; c.font = font_header; c.alignment = align_center; c.border = thin_border
ws.row_dimensions[3].height = 30

for ri, (num, dim, item, content, basis, typ) in enumerate(analysis, 4):
    ws.cell(row=ri, column=1, value=num).font = font_data; ws.cell(row=ri, column=1).alignment = align_center; ws.cell(row=ri, column=1).border = thin_border; ws.cell(row=ri, column=1).fill = fill_white
    ws.cell(row=ri, column=2, value=dim).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK); ws.cell(row=ri, column=2).alignment = align_left; ws.cell(row=ri, column=2).border = thin_border; ws.cell(row=ri, column=2).fill = fill_light
    ws.cell(row=ri, column=3, value=item).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK); ws.cell(row=ri, column=3).alignment = align_left; ws.cell(row=ri, column=3).border = thin_border; ws.cell(row=ri, column=3).fill = fill_white
    ws.cell(row=ri, column=4, value=content).font = font_data; ws.cell(row=ri, column=4).alignment = align_left; ws.cell(row=ri, column=4).border = thin_border; ws.cell(row=ri, column=4).fill = fill_white
    ws.cell(row=ri, column=5, value=basis).font = font_source; ws.cell(row=ri, column=5).alignment = align_center; ws.cell(row=ri, column=5).border = thin_border; ws.cell(row=ri, column=5).fill = fill_white
    type_cell = ws.cell(row=ri, column=6, value=typ); type_cell.font = font_data; type_cell.alignment = align_center; type_cell.border = thin_border
    color_map = {"财报事实": "C6EFCE", "分析": "FFEB9C", "推测": "FFC7CE"}
    type_cell.fill = PatternFill(start_color=color_map.get(typ, "FFFFFF"), end_color=color_map.get(typ, "FFFFFF"), fill_type="solid")
    ws.row_dimensions[ri].height = 100

ws.freeze_panes = "A4"
wb.save(OUTPUT)
print(f"[OK] Analysis Excel saved to: {OUTPUT}")
