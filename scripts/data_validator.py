# -*- coding: utf-8 -*-
"""
年报数据验证器 (Annual Report Data Validator)
annual-report-extractor 技能配套脚本

功能：
1. 完整性检查 — 22项是否全部填充
2. 数值合理性校验 — 毛利率/ROE/负债率等是否在合理范围
3. 数据来源标注检查 — 是否每项都有来源说明与来源级别
4. 规则一致性评分 — 统计来源级别标注、数据缺口说明和规则执行一致性
5. 生成验证报告

使用方法：
python data_validator.py <excel_file_path>
python data_validator.py fuyao_FY2025_22items.xlsx

Author: QClaw
Date: 2026-04-23
"""
import sys
import re
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print('[ERROR] openpyxl not found. Run: pip install openpyxl')
    sys.exit(1)


# =====================================================================
# 验证规则
# =====================================================================

NUMERIC_RULES = {
    '营收': (0, 100000, '亿元', '营业收入'),
    '净利润': (0, 10000, '亿元', '归母净利润'),
    '毛利率': (0, 100, '%', '毛利率'),
    '净利率': (0, 80, '%', '净利率'),
    'ROE': (-50, 200, '%', '净资产收益率'),
    '负债率': (0, 100, '%', '资产负债率'),
    'EPS': (-100, 1000, '元', '每股收益'),
    '每股收益': (-100, 1000, '元', '每股收益'),
    '分红': (0, 100, '元', '每股分红'),
    '分红率': (0, 200, '%', '派现率/分红率'),
    '营收增长': (-50, 200, '%', '营收同比增长率'),
    '净利润增长': (-100, 500, '%', '净利润同比增长率'),
    '市场份额': (0, 100, '%', '市场份额'),
    '供应商占比': (0, 100, '%', '供应商集中度'),
    '客户占比': (0, 100, '%', '客户集中度'),
    '研发投入占比': (0, 50, '%', '研发费用占营收比'),
}

# 22项标准清单
ITEMS_22 = [
    (1, '公司名字'),
    (2, '公司市值'),
    (3, '公司主营业务'),
    (4, '市场份额'),
    (5, '未来增长率'),
    (6, '上游供应商'),
    (7, '下游客户'),
    (8, '原材料成本'),
    (9, '资本开支'),
    (10, '行业平均毛利率'),
    (11, '公司毛利率'),
    (12, '行业平均ROE'),
    (13, '公司ROE'),
    (14, '行业平均负债率'),
    (15, '公司负债率'),
    (16, '合同负债'),
    (17, '营收增长率'),
    (18, 'PE百分位'),
    (19, 'PB百分位'),
    (20, '美股同类公司'),
    (21, '股票增减持'),
    (22, '高管增减持'),
]


# =====================================================================
# 核心验证函数
# =====================================================================

def extract_numeric(text):
    """从文本中提取数值（支持亿、万、%等）"""
    if not text or str(text).strip() in ['[待填充]', 'N/A', '', '无', '暂无']:
        return None
    text = str(text).strip()
    # 提取数字+单位
    patterns = [
        r'(-?[\d.]+)\s*亿',    # 亿
        r'(-?[\d.]+)\s*万',    # 万
        r'(-?[\d.]+)\s*%',     # %
        r'(-?[\d.]+)',          # 纯数字
    ]
    multipliers = [1e8, 1e4, 1, 1]
    for pattern, mult in zip(patterns, multipliers):
        m = re.search(pattern, text)
        if m:
            try:
                return float(m.group(1)) * mult
            except ValueError:
                pass
    return None


def extract_percent(text):
    """从文本中提取百分比（0-100范围）"""
    if not text:
        return None
    text = str(text).strip()
    # 处理 "30-34%" 格式
    m = re.search(r'(\d+\.?\d*)\s*[-–]\s*\d+\.?\d*\s*%', text)
    if m:
        return float(m.group(1))
    # 处理 "34%" 格式
    m = re.search(r'(\d+\.?\d*)\s*%', text)
    if m:
        val = float(m.group(1))
        return val if val <= 100 else val / 100 * 100
    return None


def validate_item_value(item_name, value_text):
    """验证单项数据的数值合理性"""
    issues = []
    
    if not value_text or str(value_text).strip() in ['[待填充]', 'N/A', '', '无', '暂无', '待填充']:
        return [], None
    
    for rule_key, (min_val, max_val, unit, desc) in NUMERIC_RULES.items():
        if rule_key in item_name or rule_key in str(value_text):
            if unit == '%':
                num = extract_percent(str(value_text))
            else:
                num = extract_numeric(str(value_text))
            
            if num is not None:
                if unit == '%':
                    if not (min_val <= num <= max_val):
                        issues.append(f"{desc}={num}{unit}，超出合理范围[{min_val},{max_val}]{unit}")
                else:
                    text = str(value_text)
                    if '亿' in text and num > 1e8:
                        raw = num / 1e8
                    elif '万' in text and num > 1e4:
                        raw = num / 1e4
                    else:
                        raw = num
                    if not (min_val <= raw <= max_val):
                        issues.append(f"{desc}={raw:.2f}{unit}，超出合理范围[{min_val},{max_val}]{unit}")
    
    return issues, None


def check_data_source(cell_text):
    """检查数据来源标注质量"""
    if not cell_text:
        return 'missing', '数据来源缺失'
    text = str(cell_text).strip()
    if text in ['[待填充]', 'N/A', '', '无', '暂无', '待填充']:
        return 'missing', '未填充数据来源'
    # 必须包含来源字段与来源级别
    has_source = any(kw in text for kw in ['来源级别', '来源：', '来源:', '数据来源'])
    if not has_source:
        return 'warning', '数据来源标注不明确'
    if '来源级别' not in text:
        return 'warning', '缺少来源级别标注'
    return 'ok', '有来源标注'


def check_completeness(data_rows):
    """检查22项数据完整性"""
    missing = []
    filled = []
    empty = []
    
    for seq, item_name, value, source in data_rows:
        if seq < 1 or seq > 22:
            continue
        if not value or str(value).strip() in ['[待填充]', '', '无', '暂无', '待填充', 'N/A']:
            empty.append((seq, item_name))
        else:
            filled.append((seq, item_name))
    
    return filled, empty


def classify_reliability(source_text):
    """分类数据来源级别与状态"""
    if not source_text:
        return 'unknown'
    text = str(source_text)
    if '来源级别：第1级' in text or '来源级别:第1级' in text:
        return 'level1'
    if '来源级别：第2级' in text or '来源级别:第2级' in text:
        return 'level2'
    if '来源级别：第3级' in text or '来源级别:第3级' in text:
        return 'level3'
    if '来源级别：第4级' in text or '来源级别:第4级' in text:
        return 'level4'
    if '来源级别：第5级' in text or '来源级别:第5级' in text:
        return 'level5'
    if any(kw in text for kw in ['数据缺口', '未披露', '未获取', '无法获取', '商业机密', '来源级别：未获取', '来源级别:未获取']):
        return 'gap'
    return 'unknown'


def data_rows_to_enhanced_dict(data_rows):
    """将22项Excel行转换为增强验证器字段字典"""
    field_map = {
        1: 'company_name', 2: 'market_cap', 3: 'main_business', 4: 'market_share',
        5: 'growth_forecast', 6: 'suppliers', 7: 'customers', 8: 'raw_materials',
        9: 'capex', 10: 'industry_gross_margin', 11: 'gross_margin', 12: 'industry_roe',
        13: 'roe', 14: 'industry_debt_ratio', 15: 'debt_ratio', 16: 'contract_liabilities',
        17: 'revenue_growth', 18: 'pe_percentile', 19: 'pb_percentile', 20: 'us_peers',
        21: 'share_changes', 22: 'executive_changes'
    }
    data = {}
    for row in data_rows:
        if len(row) < 4:
            continue
        seq, item_name, value, source = row[0], row[1], row[2], row[3]
        key = field_map.get(seq)
        if not key:
            continue
        value_text = '' if str(value).strip() in ['[待补充]', '待补充', '[待填充]', '待填充', 'N/A', ''] else value
        data[key] = value_text
        data[f'{key}_source'] = source
        data[f'{key}_reliability'] = source
        if seq == 17:
            data['revenue'] = value_text
            data['revenue_source'] = source
            data['revenue_reliability'] = source
    return data


def run_enhanced_validation(data_rows):
    """运行五维度增强验证,失败时降级为空结果"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)
        from data_validator_enhanced import DataValidatorEnhanced
        validator = DataValidatorEnhanced()
        enhanced_data = data_rows_to_enhanced_dict(data_rows)
        return validator.validate(enhanced_data)
    except Exception as e:
        return {
            'overall_score': 0,
            'passed': False,
            'dimensions': {},
            'issues': [f'五维度验证未执行: {e}'],
            'warnings': [],
        }

# 生成表后进行最后的交付验证函数
def build_acceptance_summary(filepath, data_rows, filled, empty, all_issues, source_issues, source_stats, enhanced_result):
    """构建可交付验收摘要"""
    file_exists = os.path.exists(filepath)
    row_count_ok = len(data_rows) == 22
    readable = True
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        wb.close()
    except Exception:
        readable = False

    validation_sheet_exists = False
    try:
        wb = openpyxl.load_workbook(filepath)
        validation_sheet_exists = '数据验证' in wb.sheetnames
        wb.close()
    except Exception:
        validation_sheet_exists = False

    unknown_count = source_stats.get('unknown', 0)
    gap_count = source_stats.get('gap', 0)
    completeness_rate = len(filled) / 22 * 100 if 22 else 0
    enhanced_score = enhanced_result.get('overall_score', 0)

    deliverable = all([
        file_exists,
        readable,
        row_count_ok,
        completeness_rate >= 100,
        len(all_issues) == 0,
        len(source_issues) == 0,
        unknown_count == 0,
        enhanced_score >= 70,
    ])

    next_steps = []
    if not file_exists:
        next_steps.append('先确认表1文件是否真实落盘')
    if not readable:
        next_steps.append('先修复 Excel 文件可读性，再继续验证')
    if not row_count_ok:
        next_steps.append('检查表1结构，确认是否完整保留 22 项数据')
    if empty:
        next_steps.append(f'补充未填充的 {len(empty)} 项数据')
    if all_issues:
        next_steps.append(f'处理 {len(all_issues)} 个数值合理性问题')
    if source_issues:
        next_steps.append(f'完善 {len(source_issues)} 项来源标注问题')
    if unknown_count > 0:
        next_steps.append('将来源不明项统一改为第1-5级或数据缺口')
    if gap_count > 0:
        next_steps.append(f'补充 {gap_count} 项数据缺口说明或继续补数')
    if enhanced_score < 70:
        next_steps.append('优先处理五维度验证中的问题，再判断是否可交付')
    if not next_steps:
        next_steps.append('当前表1已满足最小验收要求，可继续交付或生成表2联查结果')

    return {
        'file_exists': file_exists,
        'readable': readable,
        'row_count_ok': row_count_ok,
        'row_count': len(data_rows),
        'validation_sheet_exists': validation_sheet_exists,
        'deliverable': deliverable,
        'next_steps': next_steps,
    }


def validate_excel(filepath):
    """主验证函数"""
    print(f'\n{"="*60}')
    print(f'  年报数据验证器 v1.0')
    print(f'{"="*60}')
    print(f'  文件: {filepath}')
    print(f'  时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print(f'{"="*60}\n')
    
    if not os.path.exists(filepath):
        print(f'[ERROR] 文件不存在: {filepath}')
        return
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # 读取数据行（第4行起）
    data_rows = []
    for row in ws.iter_rows(min_row=4, max_row=25, values_only=True):
        if row[0] is not None:
            data_rows.append(row)
    
    print(f'[INFO] 共读取 {len(data_rows)} 行数据\n')
    
    # ---- 1. 完整性检查 ----
    print(f'[CHECK 1] 22项数据完整性')
    print(f'  {"-"*50}')
    filled, empty = check_completeness(data_rows)
    print(f'  已填充: {len(filled)}/22 项')
    print(f'  未填充: {len(empty)}/22 项')
    if empty:
        print(f'  未填充项:')
        for seq, name in empty:
            print(f'    #{seq} {name}')
    completeness_rate = len(filled) / 22 * 100
    print(f'  完整率: {completeness_rate:.1f}%\n')
    
    # ---- 2. 数值校验 ----
    print(f'[CHECK 2] 数值合理性校验')
    print(f'  {"-"*50}')
    all_issues = []
    for row in data_rows:
        if len(row) >= 3:
            seq, item_name, value = row[0], row[1], row[2]
            issues, _ = validate_item_value(str(item_name), str(value) if value else '')
            for issue in issues:
                all_issues.append(f'  #{seq} {item_name}: {issue}')
                print(f'  #{seq} {item_name}: {issue}')
    if not all_issues:
        print(f'  [PASS] 全部数值在合理范围内\n')
    else:
        print(f'  共发现 {len(all_issues)} 个潜在问题\n')
    
    # ---- 3. 数据来源检查 ----
    print(f'[CHECK 3] 数据来源标注检查')
    print(f'  {"-"*50}')
    source_stats = {'level1': 0, 'level2': 0, 'level3': 0, 'level4': 0, 'level5': 0, 'gap': 0, 'unknown': 0, 'missing': 0}
    source_issues = []
    
    for row in data_rows:
        if len(row) >= 4:
            seq, item_name, value, source = row[0], row[1], row[2], row[3]
            status, msg = check_data_source(str(source) if source else '')
            reliability = classify_reliability(str(source) if source else '')
            source_stats[reliability] += 1
            if status != 'ok':
                source_issues.append(f'  #{seq} {item_name}: {msg}')
                print(f'  #{seq} {item_name}: {msg}')
    
    if not source_issues:
        print(f'  [PASS] 所有数据项均有来源标注\n')
    print(f'  来源级别统计:')
    print(f'    第1级:       {source_stats["level1"]} 项')
    print(f'    第2级:       {source_stats["level2"]} 项')
    print(f'    第3级:       {source_stats["level3"]} 项')
    print(f'    第4级:       {source_stats["level4"]} 项')
    print(f'    第5级:       {source_stats["level5"]} 项')
    print(f'    数据缺口:     {source_stats["gap"]} 项')
    print(f'    来源不明:     {source_stats["unknown"]} 项\n')
    
    # ---- 4. 规则一致性评分 ----
    print(f'[CHECK 4] 规则一致性综合评分')
    print(f'  {"-"*50}')
    level1 = source_stats['level1']
    level2 = source_stats['level2']
    level3 = source_stats['level3']
    level4 = source_stats['level4']
    level5 = source_stats['level5']
    gap = source_stats['gap']
    unknown = source_stats['unknown']
    
    total = level1 + level2 + level3 + level4 + level5 + gap + unknown
    if total == 0:
        total = 22
    
    # 一致性评分：重点检查来源级别标注、缺口说明、未知来源比例
    score = ((level1 + level2 + level3 + level4 + level5) * 100 + gap * 70 + unknown * 30) / total
    
    if score >= 90:
        grade = 'A'
        grade_desc = '优秀'
    elif score >= 75:
        grade = 'B'
        grade_desc = '良好'
    elif score >= 60:
        grade = 'C'
        grade_desc = '及格'
    else:
        grade = 'D'
        grade_desc = '较差'
    
    known_pct = (level1 + level2 + level3 + level4 + level5) / total * 100
    unknown_pct = unknown / total * 100
    
    print(f'  综合评分: {score:.1f}/100 ({grade}级 {grade_desc})')
    print(f'  已标注来源级别占比: {known_pct:.1f}% (目标=100%)')
    print(f'  来源不明占比: {unknown_pct:.1f}% (目标=0%)')
    print(f'  数据缺口: {gap} 项')
    
    if known_pct < 100:
        print(f'  [WARN] 存在未按规则标注来源级别的数据项，需补齐来源级别')
    if unknown_pct > 0:
        print(f'  [WARN] 存在来源不明数据项，需统一改为第1-5级或数据缺口')
    if gap > 3:
        print(f'  [WARN] 数据缺口过多({gap}项)，建议继续补充')
    print()

    # ---- 5. 五维度增强验证 ----
    print(f'[CHECK 5] 五维度增强验证')
    print(f'  {"-"*50}')
    enhanced_result = run_enhanced_validation(data_rows)
    enhanced_score = enhanced_result.get('overall_score', 0)
    print(f'  五维度评分: {enhanced_score:.1f}/100')
    for dim_name, dim_data in enhanced_result.get('dimensions', {}).items():
        print(f'  {dim_name}: {dim_data.get("score", 0):.1f}')
    if enhanced_result.get('issues'):
        print(f'  增强验证问题: {len(enhanced_result.get("issues", []))} 个')
    if enhanced_result.get('warnings'):
        print(f'  增强验证警告: {len(enhanced_result.get("warnings", []))} 个')
    print()
    
    # ---- 6. 生成摘要 ----
    acceptance = build_acceptance_summary(
        filepath=filepath,
        data_rows=data_rows,
        filled=filled,
        empty=empty,
        all_issues=all_issues,
        source_issues=source_issues,
        source_stats=source_stats,
        enhanced_result=enhanced_result,
    )

    print(f'{"="*60}')
    print(f'  验证摘要')
    print(f'{"="*60}')
    print(f'  完整率:     {completeness_rate:.0f}%  ({len(filled)}/22项)')
    print(f'  数值校验:   {"PASS" if not all_issues else "WARN"}  ({len(all_issues)}个问题)')
    print(f'  来源标注:   {"PASS" if not source_issues else "WARN"}  ({len(source_issues)}个问题)')
    print(f'  规则一致性评分: {score:.1f}/100 ({grade}级)')
    print(f'  五维度评分: {enhanced_score:.1f}/100')
    print(f'  第1级:       {level1}项 | 第2级: {level2}项 | 第3级: {level3}项')
    print(f'  第4级:       {level4}项 | 第5级: {level5}项 | 数据缺口: {gap}项')
    print(f'  来源不明:     {unknown}项')
    print(f'{"="*60}\n')

    print(f'[ACCEPTANCE] 最小验收摘要:')
    print(f'  文件存在:     {"PASS" if acceptance["file_exists"] else "FAIL"}')
    print(f'  文件可回读:   {"PASS" if acceptance["readable"] else "FAIL"}')
    print(f'  表1结构22项:  {"PASS" if acceptance["row_count_ok"] else "FAIL"}  ({acceptance["row_count"]}项)')
    print(f'  数据验证Sheet: {"PASS" if acceptance["validation_sheet_exists"] else "WARN"}')
    print(f'  可继续交付:   {"YES" if acceptance["deliverable"] else "NO"}\n')

    print(f'[NEXT STEPS]')
    for idx, step in enumerate(acceptance['next_steps'], 1):
        print(f'  {idx}. {step}')

    print(f'\n[SUGGESTIONS] 改进建议:')
    if completeness_rate < 100:
        print(f'  1. 补充未填充的 {len(empty)} 项数据')
    if known_pct < 100:
        print(f'  2. 补齐缺少来源级别的数据项标注')
    if gap > 0:
        print(f'  3. 标注 {gap} 项数据缺口原因（年报未披露/商业机密等）')
    if source_issues:
        print(f'  4. 完善 {len(source_issues)} 项数据来源标注')
    if unknown_pct > 0:
        print(f'  5. 将来源不明项统一改为第1-5级或数据缺口')
    if not all_issues and completeness_rate == 100 and known_pct == 100 and unknown_pct == 0:
        print(f'  [OK] 数据质量良好，规则执行一致')
    
    wb.close()
    return {
        'completeness': completeness_rate,
        'score': score,
        'grade': grade,
        'issues': all_issues,
        'source_issues': source_issues,
        'level1_count': level1,
        'level2_count': level2,
        'level3_count': level3,
        'level4_count': level4,
        'level5_count': level5,
        'unknown_count': unknown,
        'gap_count': gap,
        'enhanced': enhanced_result,
        'enhanced_score': enhanced_score,
        'acceptance': acceptance,
    }


# =====================================================================
# 3. 生成验证报告Excel
# =====================================================================

def _populate_validation_sheet(ws, validation_result, excel_path):
    """填充验证报告Sheet"""
    TITLE_FILL = PatternFill('solid', fgColor='006B5A')
    HEADER_FILL = PatternFill('solid', fgColor='00E8F5F0')
    WARN_FILL = PatternFill('solid', fgColor='FFF3CD')
    PASS_FILL = PatternFill('solid', fgColor='E8F5E9')
    DATA_FILL = PatternFill('solid', fgColor='FFFFFF')
    THIN_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    ws.merge_cells('A1:D1')
    ws['A1'] = f'年报数据验证报告 - {os.path.basename(excel_path)}'
    ws['A1'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    summary = [
        ('完整率', f'{validation_result["completeness"]:.1f}%'),
        ('规则一致性评分', f'{validation_result["score"]:.1f}/100'),
        ('质量等级', f'{validation_result["grade"]}级'),
        ('第1级', f'{validation_result.get("level1_count", 0)}项'),
        ('第2级', f'{validation_result.get("level2_count", 0)}项'),
        ('第3级', f'{validation_result.get("level3_count", 0)}项'),
        ('第4级', f'{validation_result.get("level4_count", 0)}项'),
        ('第5级', f'{validation_result.get("level5_count", 0)}项'),
        ('数据缺口', f'{validation_result["gap_count"]}项'),
        ('数值问题', f'{len(validation_result["issues"])}个'),
        ('来源标注问题', f'{len(validation_result["source_issues"])}个'),
        ('五维度评分', f'{validation_result.get("enhanced_score", 0):.1f}/100'),
    ]

    for i, (label, value) in enumerate(summary):
        row = i + 2
        ws.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=value).font = Font(name='Arial', size=10)
        ws.cell(row=row, column=2).fill = DATA_FILL
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.row_dimensions[row].height = 22

    row = 10
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value='改进建议').font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    ws.cell(row=row, column=1).fill = TITLE_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER

    suggestions = []
    if validation_result['completeness'] < 100:
        suggestions.append('1. 补充未填充的数据项')
    if validation_result.get('unknown_count', 0) > 0:
        suggestions.append('2. 将来源不明项统一改为第1-5级或数据缺口')
    if validation_result['gap_count'] > 0:
        suggestions.append('3. 补充数据缺口说明（年报未披露/商业机密）')
    if validation_result['source_issues']:
        suggestions.append('4. 完善数据来源与来源级别标注')

    if suggestions:
        for j, s in enumerate(suggestions):
            r = row + j + 1
            ws.merge_cells(f'A{r}:D{r}')
            ws.cell(row=r, column=1, value=s).font = Font(name='Arial', size=10)
            ws.cell(row=r, column=1).fill = WARN_FILL
            ws.cell(row=r, column=1).border = THIN_BORDER
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[r].height = 22
    else:
        r = row + 1
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(row=r, column=1, value='[OK] 数据质量良好').font = Font(name='Arial', size=10)
        ws.cell(row=r, column=1).fill = PASS_FILL
        ws.cell(row=r, column=1).border = THIN_BORDER

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30


def generate_validation_report(validation_result, excel_path, output_path=None):
    """生成独立验证报告Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '验证报告'
    _populate_validation_sheet(ws, validation_result, excel_path)

    if output_path is None:
        base, ext = os.path.splitext(excel_path)
        output_path = f'{base}_验证报告.xlsx'

    wb.save(output_path)
    print(f'\n[OK] 验证报告已生成: {output_path}')


def append_validation_sheet(validation_result, excel_path):
    """将验证报告追加/更新到原数据表工作簿"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        if '数据验证' in wb.sheetnames:
            del wb['数据验证']
        ws = wb.create_sheet('数据验证')
        _populate_validation_sheet(ws, validation_result, excel_path)
        wb.save(excel_path)
        wb.close()
        print(f'[OK] 验证报告已内嵌到数据表1: {excel_path} -> Sheet 数据验证')
    except Exception as e:
        print(f'[WARN] 验证报告内嵌失败: {e}')


# =====================================================================
# 4. 入口
# =====================================================================

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python data_validator.py <excel文件路径>')
        print('示例: python data_validator.py fuyao_FY2025_22items.xlsx')
        sys.exit(1)
    
    excel_path = sys.argv[1]
    result = validate_excel(excel_path)
    
    if result:
        # 自动生成验证报告
        base, ext = os.path.splitext(excel_path)
        report_path = f'{base}_验证报告.xlsx'
        generate_validation_report(result, excel_path, report_path)
        append_validation_sheet(result, excel_path)
