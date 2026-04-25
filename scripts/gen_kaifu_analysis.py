# -*- coding: utf-8 -*-
"""Generate Kai Fu Energy (00007.HK) FY2025 Financial Analysis Report (Table 2)"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUTPUT = r"C:\Users\lishaoming\.qclaw\workspace\SKILLS\凯富能源_00007HK_FY2025_财报解读.xlsx"

GREEN_TITLE = "006B5A"; GREEN_LIGHT = "00E8F5F0"; WHITE = "FFFFFF"; BLACK = "000000"
fill_title = PatternFill(start_color=GREEN_TITLE, end_color=GREEN_TITLE, fill_type="solid")
fill_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
font_title = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Microsoft YaHei", size=10, color=BLACK)
font_source = Font(name="Microsoft YaHei", size=9, color="666666")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"), top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

analysis = [
    (1, "一、3分钟看懂版", "整体判断",
     "高风险警示。凯富能源(00007.HK)是典型港股仙股，负债率94.87%，连续多年亏损(2023年巨亏17.63亿港元)，2025年H1才扭亏为盈(净利润仅1186万港元)。公司多次更名(凯富能源->香港金融集团->凯富能源投资)，业务方向摇摆不定。投资风险极高，不建议普通投资者参与。",
     "数据项2/11/13/15/17", "分析"),

    (2, "一、3分钟看懂版", "核心原因",
     "1)2025年H1扭亏主因营收增长73.6%(商品贸易收入恢复)\n2)但毛利率仅4.95%，盈利质量极低\n3)负债率94.87%，濒临资不抵债\n4)2023年巨亏17.63亿港元(资产减值/投资损失)，元气大伤\n5)控股股东持股66.93%，流通盘极小",
     "数据项11/13/15/17", "分析"),

    (3, "二、公司靠什么赚钱", "主营业务",
     "凯富能源是一家多元化投资公司(经历多次转型)：\n1.石油天然气：美国德克萨斯州油气资产(2.33平方英里)\n2.地产：湛江综合经营项目(500亩/130万平米)+北京铜锣场\n3.金融服务：证券投资/企业融资/资产管理/保险经纪\n4.商品贸易：贵金属/大宗商品买卖\n\n当前营收主要来自商品贸易(低毛利高周转)",
     "数据项3", "财报事实"),

    (4, "二、公司靠什么赚钱", "利润来源",
     "2025年H1扭亏的利润来源：\n1.商品贸易收入增长(+73.6%)\n2.投资物业公允价值变动(如有)\n3.行政开支/融资成本控制\n\n但需注意：毛利率仅4.95%，意味着每100元收入只赚不到5元，稍有波动就可能再次亏损。",
     "数据项3/11", "分析"),

    (5, "二、公司靠什么赚钱", "经营变化",
     "最大变化：\n1.公司再次更名(香港金融集团->凯富能源投资/Wisdom Wealth)，暗示战略重心从金融回归能源\n2.2025年H1扭亏为盈，是近3年首次盈利\n3.负债率仍高达94.10%，财务结构未改善\n4.营收波动极大(2023年6.68亿->2024年2.97亿->2025年H1 3.40亿)",
     "数据项3/15/17", "财报事实"),

    (6, "三、关键财务数据解读", "营收",
     "2024年营收2.97亿港元(-56.5%)，2025年H1营收3.40亿港元(+73.6%)。看似强劲复苏，但：\n1.营收波动极大，不可持续\n2.主要来自商品贸易(低毛利业务)\n3.2024年基数极低(仅2.97亿)，73.6%增速含金量低\n4.半年3.40亿意味着全年可能6-7亿，但能否维持是问题",
     "数据项17", "分析"),

    (7, "三、关键财务数据解读", "归母净利润",
     "2024年净亏损1.08亿港元(vs 2023年巨亏17.63亿港元，亏损收窄94%)\n2025年H1净利润1,186万港元(扭亏)\n\n扭亏值得肯定，但1,186万港元利润对于一家负债率94.87%的公司而言微不足道。",
     "数据项13/17", "分析"),

    (8, "三、关键财务数据解读", "扣非净利润",
     "港股通常不区分扣非净利润。2023年巨亏17.63亿港元中，可能包含大量资产减值和投资损失(非经常性)。2024年亏损1.08亿港元已大幅收窄，说明核心经营亏损在减少。\n2025年H1扭亏是否可持续，需关注减值支出和投资收益变化。",
     "数据项17", "推测"),

    (9, "三、关键财务数据解读", "毛利率",
     "2024年毛利率6.39%，2025年H1仅4.95%。极低！\n\n毛利率低于10%意味着：\n1.公司定价能力弱，处于产业链弱势地位\n2.商品贸易业务天然低毛利(2-8%属正常)\n3.油气/地产/金融业务未能贡献高毛利\n4.稍有成本波动就可能亏损",
     "数据项11", "分析"),

    (10, "三、关键财务数据解读", "净利率",
     "2024年净利率-36.3%(亏损1.08亿/营收2.97亿)\n2025年H1净利率3.47%(净利1186万/营收3.40亿)\n\n净利率3.47%极低，意味着经营效率很差。对比：天安中国2025年净利率16.8%，恒生银行50%。",
     "数据项11/17", "分析"),

    (11, "三、关键财务数据解读", "经营现金流",
     "2025年H1每股经营现金流-0.0023港元(负值)\n2024年全年+0.0149港元(微正)\n\n虽然利润转正，但经营现金流为负，说明利润的现金含金量不足。可能原因：应收款增加/预付款增加/存货增加。",
     "数据项17", "分析"),

    (12, "三、关键财务数据解读", "资产负债率",
     "2024年负债率94.87%，2025年H1仍高达94.10%。\n\n这是最危险的信号：\n1.负债率接近95%，意味着净资产仅占总资产5-6%\n2.任何资产减值都可能导致资不抵债\n3.融资成本高(利息支出5,837万港元/2024年)\n4.偿债压力巨大，流动性风险极高",
     "数据项15", "财报事实"),

    (13, "三、关键财务数据解读", "应收账款",
     "公司业务多元化，应收账款来源复杂：\n- 商品贸易应收\n- 油气销售应收\n- 金融服务应收\n\n2024年流动比率0.95(低于1)，意味着短期资产不足以覆盖短期负债，流动性紧张。",
     "数据项3/7", "推测"),

    (14, "三、关键财务数据解读", "存货",
     "公司存货主要为：\n- 商品贸易库存(贵金属/大宗商品)\n- 油气库存\n- 开发中物业\n\n商品贸易存货受市场价格波动影响大，可能面临跌价风险。",
     "数据项8/9", "分析"),

    (15, "三、关键财务数据解读", "ROE",
     "2024年ROE -5.57%，2023年-61.51%，2025年H1仅0.64%。\n\nROE长期为负或极低，说明公司资本效率极差。即使2025年H1扭亏，0.64%的ROE远不如银行存款利率。",
     "数据项13", "分析"),

    (16, "三、关键财务数据解读", "分红/回购/融资",
     "公司连续多年亏损，无分红记录。\n2024年6月获亚洲联网科技1.39亿人民币战略投资。\n\n作为仙股，回购可能性极低。融资主要依赖控股股东(持股66.93%)和债务。",
     "数据项21/22", "推测"),

    (17, "四、识别隐藏风险", "利润质量",
     "利润质量极差：\n1.毛利率仅4.95%，盈利脆弱\n2.经营现金流为负(-0.0023港元/股)\n3.扭亏主因商品贸易收入增加，非核心业务改善\n4.1186万港元利润对94%负债率的公司微不足道\n5.任何资产减值都可能吞噬全部利润",
     "数据项11/13/17", "分析"),

    (18, "四、识别隐藏风险", "现金流恶化",
     "2025年H1经营现金流为负，尽管利润转正。原因可能是：\n1.商品贸易需要大量营运资金(进货/应收)\n2.利息支出5,837万港元/年持续消耗现金\n3.投资物业公允价值变动不影响现金流\n4.流动性紧张(流动比率0.95<1)",
     "数据项15/17", "分析"),

    (19, "四、识别隐藏风险", "资产风险",
     "最大资产风险：\n1.投资物业19.35亿港元：估值是否合理？是否有减值空间？\n2.美国德克萨斯州油气资产：产量/储量/价格波动风险\n3.湛江地产项目(500亩)：能否变现？是否有土地增值税？\n4.北京铜锣场/金钻之娱乐专营权：商业价值存疑\n\n如果投资物业减值10%(约1.9亿)，公司净资产将归零。",
     "数据项3/9", "分析"),

    (20, "四、识别隐藏风险", "一次性收益",
     "2023年巨亏17.63亿港元中可能包含大量一次性减值/投资损失。2024年亏损收窄至1.08亿港元，部分是减值减少(非经常性因素)。\n\n2025年H1扭亏是否包含一次性收益(如资产处置/公允价值变动)需关注年报附注。",
     "数据项17", "推测"),

    (21, "四、识别隐藏风险", "易忽略点",
     "1.公司3次更名(凯富能源->香港金融集团->凯富能源投资)，频繁更名=战略摇摆\n2.控股股东持股66.93%，公众流通盘仅33%，流动性极差\n3.流动比率0.95<1，短期偿债能力不足\n4.利息支出5,837万港元/年 vs 营收2.97亿港元，利息支出占营收近20%\n5.员工仅88人，管理2.97亿营收+19亿投资物业，人效存疑\n6.BPS仅0.55港元，PB约0.3-0.5倍，市场严重怀疑资产质量",
     "数据项2/3/15/21", "分析"),

    (22, "五、未来展望", "增长逻辑",
     "潜在增长点(但确定性极低)：\n1.美国油气资产若产量提升+油价上涨，可能贡献利润\n2.湛江地产项目若能销售/变现，可大幅改善财务\n3.金融服务业务恢复(曾是香港金融集团主营)\n4.商品贸易规模扩大(但毛利极低)\n\n但所有增长逻辑都被94%负债率和流动性危机笼罩。",
     "数据项3/5", "分析"),

    (23, "五、未来展望", "风险因素",
     "1.负债率94.87%，随时可能资不抵债\n2.利息支出占营收近20%，盈利被利息吞噬\n3.毛利率<5%，盈利极其脆弱\n4.控股股东高度集中，小股东利益难保障\n5.仙股流通性差，卖出困难\n6.公司多次更名/转型，治理稳定性差\n7.2023年巨亏17.63亿，资本基础已被侵蚀",
     "数据项5/11/15", "分析"),

    (24, "五、未来展望", "跟踪指标",
     "下一份报告最该跟踪：\n1.负债率变化(能否降到90%以下是关键)\n2.毛利率(能否突破10%是分水岭)\n3.经营现金流/净利润比率(利润质量)\n4.投资物业公允价值变动(是否减值)\n5.利息支出/营收比率(能否降低融资成本)\n6.湛江地产项目进展(能否变现)",
     "数据项9/11/15/17", "分析"),

    (25, "六、投资者视角结论", "长期投资者",
     "不建议参与。凯富能源是典型港股仙股：\n1.负债率94.87%，濒临资不抵债\n2.毛利率<5%，无核心竞争力\n3.连续多年亏损，2025年H1才微利\n4.频繁更名/转型=战略不清\n5.控股股东持股66.93%，小股东无话语权\n\n唯一看点：若湛江地产项目变现+油气资产利润释放，可能有翻身机会。但概率极低，风险远大于收益。",
     "数据项2/11/13/15", "分析"),

    (26, "六、投资者视角结论", "短期投资者",
     "不建议参与。仙股短期波动可能很大(消息驱动)，但：\n1.流动性极差(公众持股仅33%)\n2.买卖价差大，进出成本高\n3.信息不对称严重(小型公司披露有限)\n4.可能存在老千股风险(频繁更名/转型/高负债)\n\n如果一定要参与，仅可用极小仓位投机，设置严格止损。",
     "数据项2/18/19", "分析"),

    (27, "六、投资者视角结论", "机会 vs 风险",
     "风险远大于机会，属于'极高风险极低赔率'：\n\n风险：\n1.负债率94.87%=随时可能资不抵债\n2.毛利率<5%=无盈利护城河\n3.连续亏损=资本基础被侵蚀\n4.仙股=流动性差/信息不对称/老千风险\n\n理论机会：\n1.湛江地产变现(一次性改善财务)\n2.油气资产利润释放(若油价上涨)\n3.PB极低(0.3-0.5倍)，理论上有估值修复空间\n\n但PB低有原因——市场严重怀疑其资产质量。投资物业19.35亿估值是否真实，是最大问号。",
     "数据项2/5/11", "分析"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "凯富能源FY2025财报解读"
ws.column_dimensions["A"].width = 5.0; ws.column_dimensions["B"].width = 18.0; ws.column_dimensions["C"].width = 18.0; ws.column_dimensions["D"].width = 65.0; ws.column_dimensions["E"].width = 18.0; ws.column_dimensions["F"].width = 12.0

ws.merge_cells("A1:F1")
ws["A1"].value = "凯富能源投资有限公司(00007.HK) 2025年度财报解读分析表"
ws["A1"].fill = fill_title; ws["A1"].font = font_title; ws["A1"].alignment = align_center; ws.row_dimensions[1].height = 50

ws.merge_cells("A2:F2")
ws["A2"].value = "基于2025年H1数据(全年待公布) | 分析日期: 2026-04-25 | [高风险警示: 仙股/负债率94.87%/连续亏损]"
ws["A2"].fill = fill_light; ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="333333"); ws["A2"].alignment = align_center; ws.row_dimensions[2].height = 25

for ci, h in enumerate(["序号", "维度", "分析项", "内容", "依据", "类型"], 1):
    c = ws.cell(row=3, column=ci, value=h); c.fill = fill_title; c.font = font_header; c.alignment = align_center; c.border = thin_border
ws.row_dimensions[3].height = 30

for ri, (num, dim, item, content, basis, typ) in enumerate(analysis, 4):
    ws.cell(row=ri, column=1, value=num).font = font_data; ws.cell(row=ri, column=1).alignment = align_center; ws.cell(row=ri, column=1).border = thin_border; ws.cell(row=ri, column=1).fill = fill_white
    ws.cell(row=ri, column=2, value=dim).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK); ws.cell(row=ri, column=2).alignment = align_left; ws.cell(row=ri, column=2).border = thin_border; ws.cell(row=ri, column=2).fill = fill_light
    ws.cell(row=ri, column=3, value=item).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK); ws.cell(row=ri, column=3).alignment = align_left; ws.cell(row=ri, column=3).border = thin_border; ws.cell(row=ri, column=3).fill = fill_white
    ws.cell(row=ri, column=4, value=content).font = font_data; ws.cell(row=ri, column=4).alignment = align_left; ws.cell(row=ri, column=4).border = thin_border; ws.cell(row=ri, column=4).fill = fill_white
    ws.cell(row=ri, column=5, value=basis).font = font_source; ws.cell(row=ri, column=5).alignment = align_center; ws.cell(row=ri, column=5).border = thin_border; ws.cell(row=ri, column=5).fill = fill_white
    type_cell = ws.cell(row=ri, column=6, value=typ); type_cell.font = font_data; type_cell.alignment = align_center; type_cell.border = thin_border
    color_map = {"财报事实": "C6EFCE", "分析": "FFEB9C", "推测": "FFC7CE"}
    type_cell.fill = PatternFill(start_color=color_map.get(typ, "FFFFFF"), end_color=color_map.get(typ, "FFFFFF"), fill_type="solid")
    ws.row_dimensions[ri].height = 100

ws.freeze_panes = "A4"
wb.save(OUTPUT)
print(f"[OK] Analysis Excel saved to: {OUTPUT}")
