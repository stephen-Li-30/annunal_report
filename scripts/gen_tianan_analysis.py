# -*- coding: utf-8 -*-
"""Generate Tian An China (00028.HK) FY2025 Financial Analysis Report (Table 2)"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUTPUT = r"C:\Users\lishaoming\.qclaw\workspace\SKILLS\天安中国_00028HK_FY2025_财报解读.xlsx"

GREEN_TITLE = "006B5A"
GREEN_LIGHT = "00E8F5F0"
WHITE = "FFFFFF"
BLACK = "000000"

fill_title = PatternFill(start_color=GREEN_TITLE, end_color=GREEN_TITLE, fill_type="solid")
fill_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

font_title = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Microsoft YaHei", size=10, color=BLACK)
font_source = Font(name="Microsoft YaHei", size=9, color="666666")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# 6维度22项分析
analysis = [
    # (序号, 维度, 分析项, 内容, 依据, 类型)
    (1, "一、3分钟看懂版", "整体判断",
     "利好(短期)。天安中国2025年成功扭亏为盈,股东应占溢利17.68亿港元,营收暴增249%至104.98亿港元,毛利率从34%跳升至52%。但需注意:增长高度依赖上海天安1号二期(C区)单一项目交付,可持续性存疑。",
     "数据项2/3/11/13/17", "分析"),

    (2, "一、3分钟看懂版", "核心原因",
     "1)天安1号二期(C区)住宅集中交付,确认收入大幅增加(营收+249%)\n2)高毛利住宅项目推动毛利率从34%飙升至52%\n3)公司从亏损1.92亿港元转为盈利17.68亿港元,扭亏力度显著",
     "数据项3/11/17", "分析"),

    (3, "二、公司靠什么赚钱", "主营业务",
     "天安中国是一家老牌港资房企(1987年上市),主营物业发展(住宅/商业/办公楼开发)、物业投资(收租)、其他运营(酒店/物业管理/健康医护)。核心收入来自物业发展,占比超85%。",
     "数据项3", "财报事实"),

    (4, "二、公司靠什么赚钱", "利润来源",
     "主要利润来源是物业发展分部。2025年H1物业发展分部溢利44.18亿港元,是核心盈利引擎。物业投资提供稳定租金收入,健康医护等新业务尚在培育期。",
     "数据项3/11", "财报事实"),

    (5, "二、公司靠什么赚钱", "经营变化",
     "2025年最重大变化:上海天安1号二期(C区)年初交付,推动全年营收暴增。同时,已登记物业销售面积同比下滑55%至14.56万平方米,在建楼面面积117.82万方(同比-11%),意味着后续可交付面积可能减少。",
     "数据项3/17", "财报事实"),

    (6, "三、关键财务数据解读", "营收",
     "2025年营收104.98亿港元(+249%)。看似强劲,实则高度依赖天安1号二期单一项目交付。2024年营收仅30.12亿港元(基数极低),249%增速的含金量需打折。房地产结算有周期性,2026年营收可能大幅回落。",
     "数据项17", "分析"),

    (7, "三、关键财务数据解读", "归母净利润",
     "2025年股东应占溢利17.68亿港元(vs 2024年亏损1.92亿港元)。扭亏为盈值得肯定,但需注意:利润高度集中在上半年(H1利润23.45亿港元,下半年利润可能回落),且由单一项目驱动。",
     "数据项13/17", "分析"),

    (8, "三、关键财务数据解读", "扣非净利润",
     "港股通常不区分扣非净利润。公司溢利17.1亿港元,股东应占溢利17.68亿港元,差额可能来自联营公司收益等。需关注非经常性损益对利润的影响。",
     "数据项17", "推测"),

    (9, "三、关键财务数据解读", "毛利率",
     "2025年毛利率51.72%(vs 2024年34.28%)。远高于行业平均(20-30%),主因天安1号二期是上海高端住宅项目,地价成本较低,毛利率天然较高。但如此高的毛利率不可持续,回归30-35%正常水平是大概率事件。",
     "数据项11", "分析"),

    (10, "三、关键财务数据解读", "净利率",
     "2025年净利率约16.8%(溢利17.1亿/营收104.98亿)。对于房地产企业而言属中等偏上水平。但需注意,2025年毛利率异常高(52%),未来净利率将随毛利率回落而下降。",
     "数据项11/17", "分析"),

    (11, "三、关键财务数据解读", "经营现金流",
     "2025年每股经营现金流0.082港元(vs 2024年2.675港元)。经营现金流大幅下降,尽管利润转正,但现金流入并未同步改善,需关注应收款项和预付款项的变化。房地产行业现金流滞后于利润确认是常态。",
     "数据项17", "分析"),

    (12, "三、关键财务数据解读", "资产负债率",
     "2025年负债率45.81%(vs 2024年51.78%)。下降主因利润转正后净资产增加。在港股内房股中属偏低水平(行业平均65-80%),财务杠杆相对保守,偿债压力不大。",
     "数据项15", "财报事实"),

    (13, "三、关键财务数据解读", "应收账款",
     "房地产企业主要应收为购房款和租金。天安中国以住宅销售为主(客户为个人购房者),应收账款风险相对可控。但需关注物业管理费和商业租金的回收情况。",
     "数据项3/7", "推测"),

    (14, "三、关键财务数据解读", "存货",
     "房地产企业最大存货是开发中物业和待售物业。2025年在建楼面117.82万平方米(-11%),已登记销售14.56万平方米(-55%)。销售面积大幅下滑是值得警惕的信号,去化压力增大。",
     "数据项7/9", "分析"),

    (15, "三、关键财务数据解读", "ROE",
     "2025年ROE 6.42%(vs 2024年-0.77%)。ROE转正,但6.42%的水平对房地产企业而言偏低(好年份通常10-15%)。PB约0.3倍(大幅破净),市场对公司持续盈利能力存疑。",
     "数据项13", "分析"),

    (16, "三、关键财务数据解读", "分红/回购/融资",
     "天安中国作为港股小盘内房股,分红政策需查阅年报。2025年扭亏为盈后,如果恢复分红将是积极信号。此前亏损年分红可能性较低。港股回购需关注公告。",
     "数据项21/22", "推测"),

    (17, "四、识别隐藏风险", "利润质量",
     "利润质量不高。2025年利润高度依赖天安1号二期单一项目,属于'结算节奏'而非'经营趋势改善'。经营现金流仅0.082港元/股,远低于1.09港元/股的EPS,利润的现金含金量不足。",
     "数据项11/17", "分析"),

    (18, "四、识别隐藏风险", "现金流恶化",
     "2025年每股经营现金流0.082港元(vs 2024年2.675港元),大幅下降。虽然利润转正,但现金并未同步流入,可能是:1)大量资金沉淀在开发项目 2)预售款(合约负债)减少 3)应收增加。需关注后续现金流改善情况。",
     "数据项16/17", "分析"),

    (19, "四、识别隐藏风险", "资产风险",
     "房地产企业最大风险是存货减值(开发中物业贬值)。天安中国在长三角/大湾区布局,核心城市房价相对坚挺,减值风险可控。但需关注:1)商业地产空置率上升 2)三线城市项目去化困难 3)联营公司业绩波动。",
     "数据项3/9", "分析"),

    (20, "四、识别隐藏风险", "一次性收益",
     "2025年业绩改善主要来自天安1号二期交付确认,这属于正常业务结算而非一次性收益。但需注意:这种高毛利项目交付是不可持续的,后续年份毛利率大概率回归30-35%水平。",
     "数据项3/11", "分析"),

    (21, "四、识别隐藏风险", "易忽略点",
     "1)已登记物业销售面积同比-55%,是未来营收的前瞻指标,预示2026年营收可能大幅下滑\n2)在建楼面面积-11%,开发节奏放缓\n3)租金收入同比-1.6%,商业地产面临压力\n4)控股股东联合集团(00373.HK)对天安的持股稳定,但需关注关联交易\n5)PB仅0.3倍,市场对盈利持续性高度怀疑",
     "数据项3/7/9", "分析"),

    (22, "五、未来展望", "增长逻辑",
     "1)天安1号后续期数交付(如有)可提供收入支撑\n2)天安数码城/天安云谷产业园是差异化竞争优势,产业地产需求相对稳定\n3)长三角/大湾区核心城市需求韧性较强\n4)健康医护/养老是长期增长赛道\n5)低负债率为逆周期拿地提供空间",
     "数据项3/5", "分析"),

    (23, "五、未来展望", "风险因素",
     "1)中国房地产政策不确定性,行业仍在探底\n2)销售面积-55%,未来结算收入可能大幅下滑\n3)毛利率不可持续,52%大概率回落至30-35%\n4)经营现金流弱,利润质量待提升\n5)港股小盘股流动性差,估值修复困难",
     "数据项5/7/11", "分析"),

    (24, "五、未来展望", "跟踪指标",
     "下一份中期/季度报告最该跟踪:\n1)合约负债(预售款)变化 - 预示未来结算收入\n2)新增土地储备 - 判断公司扩张意愿\n3)销售面积/金额同比 - 前瞻性指标\n4)经营现金流/净利润比率 - 利润质量\n5)租金收入趋势 - 商业地产景气度",
     "数据项7/9/16", "分析"),

    (25, "六、投资者视角结论", "长期投资者",
     "天安中国PB仅0.3倍,大幅破净,如果公司能持续盈利,估值修复空间巨大。但核心问题是:2025年业绩由单一项目驱动,2026-2027年能否维持盈利是关键。长期投资者应重点关注:1)后续项目交付节奏 2)产业地产(天安数码城)的稳定收入 3)分红政策恢复。在确认盈利可持续之前,建议观望。",
     "数据项2/11/13", "分析"),

    (26, "六、投资者视角结论", "短期投资者",
     "短期来看,2025年扭亏为盈是催化事件,但股价已部分反映。PB=0.3倍的极低估值提供安全边际,但港股小盘股流动性差,短期上涨空间受制于成交量和市场情绪。关注点:1)年报发布后的分红决议 2)下一个项目交付时间 3)行业政策变化。",
     "数据项2/18/19", "分析"),

    (27, "六、投资者视角结论", "机会 vs 风险",
     "更像是'高风险高赔率'的机会。天安中国在0.3倍PB交易,如果盈利可持续,估值修复空间可达3倍以上。但2025年业绩是一次性项目驱动的'虚高',2026年大概率回落。真正的机会在于:1)产业地产(天安数码城)提供稳定现金流 2)低负债率+破净=安全边际较高。风险在于:1)销售面积持续下滑 2)行业下行周期可能持续2-3年。",
     "数据项2/5/11", "分析"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "天安中国FY2025财报解读"

ws.column_dimensions["A"].width = 5.0
ws.column_dimensions["B"].width = 18.0
ws.column_dimensions["C"].width = 18.0
ws.column_dimensions["D"].width = 65.0
ws.column_dimensions["E"].width = 18.0
ws.column_dimensions["F"].width = 12.0

# Title
ws.merge_cells("A1:F1")
ws["A1"].value = "天安中国投资有限公司(00028.HK) 2025年度财报解读分析表"
ws["A1"].fill = fill_title
ws["A1"].font = font_title
ws["A1"].alignment = align_center
ws.row_dimensions[1].height = 50

# Subtitle
ws.merge_cells("A2:F2")
ws["A2"].value = "基于2025年度经审核业绩公告(2026-03-20发布) | 分析日期: 2026-04-25"
ws["A2"].fill = fill_light
ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="333333")
ws["A2"].alignment = align_center
ws.row_dimensions[2].height = 25

# Header
headers = ["序号", "维度", "分析项", "内容", "依据", "类型"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.fill = fill_title
    c.font = font_header
    c.alignment = align_center
    c.border = thin_border
ws.row_dimensions[3].height = 30

# Data
for ri, (num, dim, item, content, basis, typ) in enumerate(analysis, 4):
    ws.cell(row=ri, column=1, value=num).font = font_data
    ws.cell(row=ri, column=1).alignment = align_center
    ws.cell(row=ri, column=1).border = thin_border
    ws.cell(row=ri, column=1).fill = fill_white

    ws.cell(row=ri, column=2, value=dim).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK)
    ws.cell(row=ri, column=2).alignment = align_left
    ws.cell(row=ri, column=2).border = thin_border
    ws.cell(row=ri, column=2).fill = fill_light

    ws.cell(row=ri, column=3, value=item).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK)
    ws.cell(row=ri, column=3).alignment = align_left
    ws.cell(row=ri, column=3).border = thin_border
    ws.cell(row=ri, column=3).fill = fill_white

    ws.cell(row=ri, column=4, value=content).font = font_data
    ws.cell(row=ri, column=4).alignment = align_left
    ws.cell(row=ri, column=4).border = thin_border
    ws.cell(row=ri, column=4).fill = fill_white

    ws.cell(row=ri, column=5, value=basis).font = font_source
    ws.cell(row=ri, column=5).alignment = align_center
    ws.cell(row=ri, column=5).border = thin_border
    ws.cell(row=ri, column=5).fill = fill_white

    type_cell = ws.cell(row=ri, column=6, value=typ)
    type_cell.font = font_data
    type_cell.alignment = align_center
    type_cell.border = thin_border
    if typ == "财报事实":
        type_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif typ == "分析":
        type_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    elif typ == "推测":
        type_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.row_dimensions[ri].height = 100

ws.freeze_panes = "A4"
wb.save(OUTPUT)
print(f"[OK] Analysis Excel saved to: {OUTPUT}")
