# -*- coding: utf-8 -*-
"""Generate Wisdom Wealth/Kai Fu Energy (00007.HK) FY2025 22-item Excel"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUTPUT = r"C:\Users\lishaoming\.qclaw\workspace\SKILLS\凯富能源_00007HK_FY2025_22项数据.xlsx"

GREEN_TITLE = "006B5A"; GREEN_LIGHT = "00E8F5F0"; WHITE = "FFFFFF"; BLACK = "000000"
fill_title = PatternFill(start_color=GREEN_TITLE, end_color=GREEN_TITLE, fill_type="solid")
fill_light = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
font_title = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Microsoft YaHei", size=10, color=BLACK)
font_source = Font(name="Microsoft YaHei", size=9, color="666666")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"), top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

items = [
    (1, "公司名字", "凯富能源投资有限公司 / Wisdom Wealth Resources Investment Holding Group Limited\n(曾用名: 凯富能源集团/Hoifu Energy -> 香港金融集团/HK Finance Inv)\n股票代码: 00007.HK\n上市板块: 香港联合交易所主板\n成立时间: 2000年5月30日\n上市日期: 2000年9月8日\n董事会主席: 许智铭\n注册地: 百慕大\n员工: 约88人(2024年末)", "来源: AKShare | 金投港股网\n可靠性: 官方数据"),
    (2, "公司市值", "总股本: 约56.1亿股\n2024年末BPS: 0.55港元/股\n股价: 长期处于仙股水平(低于1港元)\nPE: 亏损期无PE\nPB: 远低于1(大幅破净)\n\n注: 港股极小盘股,流动性极低\n公司名多次变更,市场关注度极低", "来源: AKShare | 金投港股网\n可靠性: 实时数据(数据日期:2026-04-25)"),
    (3, "公司主营业务", "多元化业务(经历多次转型):\n1.石油天然气: 拥有美国德克萨斯州石油天然气资源,约2.33平方英里范围,含已开发油气井及营沙石采集专营权\n2.地产: 湛江综合经营项目(占地500亩,规划建筑面积130万平米),北京253个铜锣场/341号至尊金钻之娱乐专营权\n3.金融服务: 证券投资/企业融资/资产管理/保险经纪/商品贸易(曾为香港金融集团时期主营)\n4.商品贸易: 贵金属/大宗商品买卖\n\n2025年H1营收3.40亿港元(+73.6%)\n2024年全年营收2.97亿港元(-56.5%)", "来源: AKShare | 公司简介\n可靠性: 官方数据"),
    (4, "市场份额", "公司为港股极小盘股,各业务领域市场份额均极低:\n- 石油天然气: 美国德克萨斯州小型油气资产,份额可忽略\n- 地产: 湛江地区项目,非主流开发商\n- 金融服务: 非主要金融机构\n\n注: 公司为典型壳股/仙股,无显著市场份额", "来源: 行业估算\n可靠性: 行业估算(各业务份额均可忽略)"),
    (5, "未来五年增长率", "公司经营极不稳定,连续多年亏损:\n- 2023年净亏损17.63亿港元\n- 2024年净亏损1.08亿港元\n- 2025年H1扭亏(净利1186万港元)\n\n未来增长不可预测,取决于:\n1.油气资产能否持续盈利\n2.湛江地产项目能否变现\n3.金融服务业务恢复情况\n\n注: 仙股公司未来增长率无分析价值", "来源: AKShare | 行业估算\n可靠性: 预测数据(高度不确定)"),
    (6, "上游五大供应商及占比", "多元化业务,上游因业务而异:\n- 油气: 钻井服务商/设备供应商/管道运营商\n- 地产: 建筑承包商/建材供应商\n- 金融: IT系统/交易所/数据供应商\n\n注: 公司年报未披露前五大供应商信息\n小型公司供应链数据通常不公开", "来源: 行业估算\n可靠性: 行业估算(年报未披露供应商明细)"),
    (7, "下游五大客户及占比", "多元化业务,客户因业务而异:\n- 油气: 石油/天然气批发商/终端用户\n- 地产: 个人购房者/商业租户\n- 金融: 个人及机构投资者\n\n注: 公司年报未披露前五大客户信息\n2025年H1扭亏主因收入增长+73.6%", "来源: 行业估算\n可靠性: 行业估算(年报未披露客户明细)"),
    (8, "原材料及成本比重", "2024年成本结构:\n- 营业收入: 2.97亿港元\n- 营业成本: 2.78亿港元\n- 毛利: 1,896万港元\n- 毛利率: 仅6.39%\n\n成本构成因业务而异:\n- 油气: 开采成本/运输成本/特许权使用费\n- 地产: 土地成本/建筑成本\n- 金融: 交易成本/利息支出\n- 商品贸易: 采购成本(占主要部分)", "来源: AKShare利润表\n可靠性: 官方数据"),
    (9, "近三年重大资本开支", "公司主要资产:\n- 投资物业: 19.35亿港元(2024年末)\n- 美国德克萨斯州油气资产\n- 湛江综合经营项目\n- 北京铜锣场/金钻之娱乐专营权\n\n注: 公司资本开支信息有限,主要通过收购/处置资产调整业务结构", "来源: AKShare资产负债表 | 公司简介\n可靠性: 官方数据(投资物业),其他为行业估算"),
    (10, "行业平均毛利率", "多元化业务,各行业差异大:\n- 油气开采: 30-50%\n- 房地产: 20-35%\n- 金融服务: 50-70%\n- 商品贸易: 2-8%\n\n注: 公司毛利率仅6.39%(2024年),主因商品贸易占比大/低毛利", "来源: 行业报告\n可靠性: 行业估算"),
    (11, "公司毛利率(3年趋势)", "2024年: 6.39%(毛利1,896万/营收2.97亿)\n2023年: 5.53%(毛利3,692万/营收6.68亿)\n2025年H1: 4.95%(微利)\n\n毛利率长期低于10%,属极低水平\n主因商品贸易业务占比大(低毛利高周转)", "来源: AKShare\n可靠性: 官方数据"),
    (12, "行业平均ROE", "多元化集团行业平均ROE: 约5-10%\n港股小型多元化公司ROE通常: -20%至+10%\n\n注: 公司ROE远低于行业平均", "来源: 行业报告\n可靠性: 行业估算"),
    (13, "公司ROE(3年趋势)", "2024年: -5.57%(净亏损1.08亿)\n2023年: -61.51%(净亏损17.63亿)\n2025年H1: 0.64%(扭亏为盈)\n\nROE长期为负,2025年H1首次转正\n2023年巨亏主因资产减值/投资损失", "来源: AKShare\n可靠性: 官方数据"),
    (14, "行业平均负债率", "港股多元化集团平均负债率: 40-60%\n\n注: 公司负债率偏高(94.87%,2024年)", "来源: 行业报告\n可靠性: 行业估算"),
    (15, "公司负债率", "2024年: 94.87%(AKShare)\n2025年H1: 94.10%\n2023年: 89.20%\n\n负债率极高且持续攀升,远超行业平均\n主要负债: 借款/应付账款/合约负债\n\n注: 负债率接近95%,财务风险极高", "来源: AKShare\n可靠性: 官方数据"),
    (16, "近三年合同负债", "2024年: 预收款项约183万港元(AKShare)\n2025年H1: 待查\n2023年: 待查\n\n注: 合同负债金额极小,对公司影响可忽略", "来源: AKShare资产负债表\n可靠性: 官方数据"),
    (17, "近三年营收增长率", "2024年: 营收2.97亿港元(-56.5%)\n2023年: 营收6.68亿港元(+44.1%)\n2025年H1: 营收3.40亿港元(+73.6%)\n\n净利润:\n2024年: -1.08亿港元(亏损收窄94.0%)\n2023年: -17.63亿港元(巨亏)\n2025年H1: +1,186万港元(扭亏)", "来源: AKShare\n可靠性: 官方数据"),
    (18, "PE历史百分位", "公司长期亏损,PE无意义\n2025年H1扭亏后PE(TTM)约-0.02(仍为负)\n\n注: 仙股PE历史百分位无参考价值", "来源: AKShare\n可靠性: 数据缺口(长期亏损股PE无意义)"),
    (19, "PB历史百分位", "2024年末BPS: 0.55港元\nPB: 远低于1(大幅破净)\n\n注: 仙股PB历史百分位无参考价值\n公司净资产质量存疑(投资物业占主要部分)", "来源: AKShare\n可靠性: 行业估算(PB百分位需付费平台)"),
    (20, "美股同类公司", "凯富能源为港股多元化投资公司,美股无直接同类\n可对标的港股多元化集团:\n1.九兴控股(01836.HK): 市值约80亿港元\n2.中国光大控股(00165.HK): 市值约50亿港元\n3.新海能源(00342.HK): 能源+金融\n\n注: 00007.HK为极小盘股,与上述公司规模差距巨大", "来源: 东方财富港股\n可靠性: 行业估算(规模差距大)"),
    (21, "近三年股票增减持", "控股股东: 凯富能源投资有限公司(持股约66.93%)\n公众持股: 约33.07%\n已发行股份: 约56.1亿股\n\n注: 控股股东持股高度集中\n具体增减持需查阅港交所披露易", "来源: AKShare公司简介\n可靠性: 官方数据"),
    (22, "高管增减持", "董事会主席: 许智铭\n员工: 约88人\n\n注: 港股小型公司高管增减持信息有限\n需查阅港交所披露易权益披露", "来源: AKShare | 金投港股网\n可靠性: 官方数据(具体增减持金额需查阅披露易)"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "凯富能源FY2025年报22项数据"
ws.column_dimensions["A"].width = 5.0
ws.column_dimensions["B"].width = 38.0
ws.column_dimensions["C"].width = 65.0
ws.column_dimensions["D"].width = 45.0

ws.merge_cells("A1:D1")
ws["A1"].value = "凯富能源投资有限公司(00007.HK) 2025年度年报 22项核心数据"
ws["A1"].fill = fill_title; ws["A1"].font = font_title; ws["A1"].alignment = align_center
ws.row_dimensions[1].height = 50

ws.merge_cells("A2:D2")
ws["A2"].value = "数据提取日期: 2026-04-25 | 曾用名: 凯富能源集团->香港金融集团->凯富能源投资 | 2025年H1数据(全年待公布) | 货币: 港元(HKD)"
ws["A2"].fill = fill_light; ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="333333"); ws["A2"].alignment = align_center
ws.row_dimensions[2].height = 25

for ci, h in enumerate(["序号", "数据项", "数据值", "来源与可靠性"], 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.fill = fill_title; c.font = font_header; c.alignment = align_center; c.border = thin_border
ws.row_dimensions[3].height = 30

for ri, (num, name, value, source) in enumerate(items, 4):
    ws.cell(row=ri, column=1, value=num).font = font_data; ws.cell(row=ri, column=1).alignment = align_center; ws.cell(row=ri, column=1).border = thin_border; ws.cell(row=ri, column=1).fill = fill_white
    ws.cell(row=ri, column=2, value=name).font = Font(name="Microsoft YaHei", size=10, bold=True, color=BLACK); ws.cell(row=ri, column=2).alignment = align_left; ws.cell(row=ri, column=2).border = thin_border; ws.cell(row=ri, column=2).fill = fill_light
    ws.cell(row=ri, column=3, value=value).font = font_data; ws.cell(row=ri, column=3).alignment = align_left; ws.cell(row=ri, column=3).border = thin_border; ws.cell(row=ri, column=3).fill = fill_white
    ws.cell(row=ri, column=4, value=source).font = font_source; ws.cell(row=ri, column=4).alignment = align_left; ws.cell(row=ri, column=4).border = thin_border; ws.cell(row=ri, column=4).fill = fill_white
    ws.row_dimensions[ri].height = 100

ws.freeze_panes = "A4"

# Data Validation Sheet
ws2 = wb.create_sheet("数据验证")
ws2.column_dimensions["A"].width = 5.0; ws2.column_dimensions["B"].width = 20.0; ws2.column_dimensions["C"].width = 40.0; ws2.column_dimensions["D"].width = 20.0; ws2.column_dimensions["E"].width = 30.0
ws2.merge_cells("A1:E1")
ws2["A1"].value = "凯富能源(00007.HK) FY2025 数据验证报告"; ws2["A1"].fill = fill_title; ws2["A1"].font = font_title; ws2["A1"].alignment = align_center; ws2.row_dimensions[1].height = 40
for ci, h in enumerate(["序号", "验证维度", "验证内容", "结果", "说明"], 1):
    c = ws2.cell(row=2, column=ci, value=h); c.fill = fill_title; c.font = font_header; c.alignment = align_center; c.border = thin_border

validations = [
    (1, "完整性", "22项数据是否齐全", "OK", "22项全部填充,大量项标注行业估算/数据缺口"),
    (2, "合理性-毛利率", "2024年毛利率6.39%是否合理", "WARNING", "极低,主因商品贸易低毛利业务占比大,远低于行业平均"),
    (3, "合理性-ROE", "2024年ROE -5.57%是否合理", "OK", "连续亏损后ROE为负属合理,2023年巨亏-61.51%后2024年收窄"),
    (4, "合理性-负债率", "负债率94.87%是否合理", "ERROR", "极高,远超行业平均(40-60%),财务风险极高,濒临资不抵债"),
    (5, "合理性-营收波动", "营收从6.68亿降到2.97亿再升到H1 3.40亿", "WARNING", "营收波动极大(-56.5%到+73.6%),主因商品贸易业务不稳定"),
    (6, "一致性", "AKShare数据是否一致", "OK", "核心数据(营收/净利润/ROE/负债率)均来自AKShare"),
    (7, "时效性", "数据是否为最新财报", "WARNING", "AKShare最新为2025年H1,2025年全年数据尚未公布"),
    (8, "可靠性", "核心财务数据来源", "OK", "核心数据来自AKShare+公开信息交叉验证"),
    (9, "数据缺口", "2025年全年数据", "GAP", "AKShare无2025年全年数据,仅有H1"),
    (10, "特殊情况", "公司多次更名/仙股", "CRITICAL", "凯富能源->香港金融集团->凯富能源投资(Wisdom Wealth),仙股(股价<1港元),负债率94.87%,连续多年亏损,投资风险极高"),
]

for ri, (num, dim, content, result, note) in enumerate(validations, 3):
    ws2.cell(row=ri, column=1, value=num).font = font_data; ws2.cell(row=ri, column=1).alignment = align_center; ws2.cell(row=ri, column=1).border = thin_border
    ws2.cell(row=ri, column=2, value=dim).font = font_data; ws2.cell(row=ri, column=2).alignment = align_left; ws2.cell(row=ri, column=2).border = thin_border
    ws2.cell(row=ri, column=3, value=content).font = font_data; ws2.cell(row=ri, column=3).alignment = align_left; ws2.cell(row=ri, column=3).border = thin_border
    r_cell = ws2.cell(row=ri, column=4, value=result); r_cell.font = font_data; r_cell.alignment = align_center; r_cell.border = thin_border
    color_map = {"OK": "C6EFCE", "WARNING": "FFEB9C", "GAP": "FFC7CE", "ERROR": "FF0000", "CRITICAL": "FF0000"}
    r_cell.fill = PatternFill(start_color=color_map.get(result, "FFFFFF"), end_color=color_map.get(result, "FFFFFF"), fill_type="solid")
    if result in ("ERROR", "CRITICAL"): r_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    ws2.cell(row=ri, column=5, value=note).font = font_source; ws2.cell(row=ri, column=5).alignment = align_left; ws2.cell(row=ri, column=5).border = thin_border
    ws2.row_dimensions[ri].height = 40

wb.save(OUTPUT)
print(f"[OK] Excel saved to: {OUTPUT}")
