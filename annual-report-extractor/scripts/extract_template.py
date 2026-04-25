#!/usr/bin/env python3
"""
年报21项数据提取模板脚本
Annual Report 21 Items Data Extraction Template

使用方法:
1. 填充 companies 列表中的公司信息
2. 运行脚本生成Excel框架
3. 手工补充搜索数据

Author: QClaw
Date: 2026-04-20
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_annual_report_excel(company_name, stock_code, fiscal_year, data_rows):
    """
    创建年报21项数据Excel文件
    
    Args:
        company_name: 公司名称（如：申洲国际）
        stock_code: 股票代码（如：2313.HK）
        fiscal_year: 财务年度（如：2025）
        data_rows: 21行数据列表，每行格式：(序号, 数据项, 具体值, 数据来源)
    
    Returns:
        生成的Excel文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{company_name} FY{fiscal_year} 21项数据'
    
    # 样式定义
    FILL1 = PatternFill('solid', fgColor='004C99')
    FILL2 = PatternFill('solid', fgColor='0066CC')
    FILL3 = PatternFill('solid', fgColor='E8F0FF')
    FILL4 = PatternFill('solid', fgColor='FFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    B = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 列宽
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 45
    
    # 标题行
    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value=f'{company_name}（{stock_code}）— FY{fiscal_year}年报 21项核心数据')
    title_cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    title_cell.fill = FILL1
    title_cell.alignment = C
    ws.row_dimensions[1].height = 36
    
    # 数据来源行
    ws.merge_cells('A2:D2')
    source_cell = ws.cell(row=2, column=1, value=f'数据来源：{company_name}{fiscal_year}年年度报告（{fiscal_year+1}年发布）| 货币单位：详见各数据项')
    source_cell.font = Font(name='Arial', italic=True, color='595959', size=9)
    source_cell.fill = PatternFill('solid', fgColor='D4E4FF')
    source_cell.alignment = C
    ws.row_dimensions[2].height = 20
    
    # 表头
    headers = ['序号', '数据项', '具体值', '数据来源/说明']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cell.fill = FILL2
        cell.alignment = C
        cell.border = B
    ws.row_dimensions[3].height = 26
    
    # 数据行
    for i, row in enumerate(data_rows):
        seq, item_name, value, source = row
        row_num = i + 4
        fill = FILL3 if row_num % 2 == 0 else FILL4
        
        for col, val in enumerate([seq, item_name, value, source], 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if col == 1:
                cell.font = Font(name='Arial', size=10)
                cell.alignment = C
            elif col == 2:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = L
            elif col == 3:
                cell.font = Font(name='Arial', size=10)
                cell.alignment = L
            else:
                cell.font = Font(name='Arial', italic=True, size=9, color='595959')
                cell.alignment = L
            cell.fill = fill
            cell.border = B
        ws.row_dimensions[row_num].height = 72
    
    # 说明行
    note_row = 4 + len(data_rows) + 1
    ws.merge_cells(f'A{note_row}:D{note_row}')
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.value = f'说明：本表数据综合整理自{company_name}{fiscal_year}年年度报告及各财经媒体报道。标注（估算）的数据为基于公开信息的合理推断。投资决策需谨慎，请以公司正式公告为准。'
    note_cell.font = Font(name='Arial', size=9, italic=True, color='7F7F7F')
    note_cell.fill = PatternFill('solid', fgColor='FFF3CD')
    note_cell.alignment = L
    note_cell.border = B
    ws.row_dimensions[note_row].height = 45
    
    # 冻结窗格
    ws.freeze_panes = 'A4'
    
    # 保存文件
    filename = f'{company_name}_FY{fiscal_year}_21items.xlsx'
    filepath = filename  # 当前目录
    wb.save(filepath)
    return filepath


# 21项数据模板
TEMPLATE_ROWS = [
    (1, '公司名字', '[待填充]', '[待填充]'),
    (2, '公司市值', '[待填充]', '[待填充]'),
    (3, '公司主营业务', '[待填充]', '[待填充]'),
    (4, '公司主营业务占世界市场份额', '[待填充]', '[待填充]'),
    (5, '预计主营业务未来五年年化增长率', '[待填充]', '[待填充]'),
    (6, '公司上游最大5家公司及占比', '[待填充]', '[待填充]'),
    (7, '公司下游五家公司及占比', '[待填充]', '[待填充]'),
    (8, '原材料是什么及占成本比重', '[待填充]', '[待填充]'),
    (9, '公司近三年重大资本开支事项', '[待填充]', '[待填充]'),
    (10, '所在行业平均毛利率', '[待填充]', '[待填充]'),
    (11, '公司毛利率', '[待填充]', '[待填充]'),
    (12, '所在行业平均净资产收益率', '[待填充]', '[待填充]'),
    (13, '公司净资产收益率', '[待填充]', '[待填充]'),
    (14, '所在行业平均负债率', '[待填充]', '[待填充]'),
    (15, '公司负债率', '[待填充]', '[待填充]'),
    (16, '过去三年营收增长率', '[待填充]', '[待填充]'),
    (17, 'PE历史百分位', '[待填充]', '[待填充]'),
    (18, 'PB历史百分位', '[待填充]', '[待填充]'),
    (19, '美股同类公司', '[待填充]', '[待填充]'),
    (20, '近三年公司股票增减持信息', '[待填充]', '[待填充]'),
    (21, '高管增减持信息', '[待填充]', '[待填充]'),
]


if __name__ == '__main__':
    # 示例用法
    filepath = create_annual_report_excel(
        company_name='示例公司',
        stock_code='123456.SH',
        fiscal_year=2025,
        data_rows=TEMPLATE_ROWS
    )
    print(f'Template created: {filepath}')
