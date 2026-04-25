# -*- coding: utf-8 -*-
"""
财报解读分析表生成器 (数据表2)
基于数据表1的22项数据，生成6维度27项分析报告

使用方法:
  python generate_analysis_report.py --input=年报数据.xlsx --output=财报解读.xlsx
  python generate_analysis_report.py --help

依赖:
  - openpyxl: Excel读写与样式
  - pandas: 数据处理

Author: QClaw
Date: 2026-04-24
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import os
from datetime import datetime

# =====================================================================
# 样式常量（固定，不得更改）
# =====================================================================
TITLE_FILL = PatternFill('solid', fgColor='006B5A')
HEADER_FILL = PatternFill('solid', fgColor='00E8F5F0')
DATA_FILL = PatternFill('solid', fgColor='FFFFFF')
ALT_FILL = PatternFill('solid', fgColor='F5FAFA')
TITLE_FONT = Font(name='Arial', size=12, bold=True, color='FFFFFF')
HEADER_FONT = Font(name='Arial', size=10, bold=False, color='000000')
DATA_FONT = Font(name='Arial', size=10, bold=False, color='000000')
DATE_FONT = Font(name='Arial', size=9, color='666666')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# =====================================================================
# 6维度27项分析框架（固定模板）
# =====================================================================
ANALYSIS_TEMPLATE = [
    # 一、3分钟看懂版 (2项)
    {"seq": 1, "dim": "一、3分钟看懂版", "item": "整体判断", "content": "", "evidence": "", "type": "分析"},
    {"seq": 2, "dim": "一、3分钟看懂版", "item": "核心原因", "content": "", "evidence": "", "type": "财报事实"},
    
    # 二、公司靠什么赚钱 (3项)
    {"seq": 3, "dim": "二、公司靠什么赚钱", "item": "主营业务", "content": "", "evidence": "数据项3", "type": "财报事实"},
    {"seq": 4, "dim": "二、公司靠什么赚钱", "item": "利润来源", "content": "", "evidence": "数据项3,11", "type": "财报事实"},
    {"seq": 5, "dim": "二、公司靠什么赚钱", "item": "经营变化", "content": "", "evidence": "数据项3", "type": "财报事实"},
    
    # 三、关键财务数据解读 (11项)
    {"seq": 6, "dim": "三、关键财务数据解读", "item": "营收", "content": "", "evidence": "数据项17", "type": "财报事实"},
    {"seq": 7, "dim": "三、关键财务数据解读", "item": "归母净利润", "content": "", "evidence": "数据项17", "type": "财报事实"},
    {"seq": 8, "dim": "三、关键财务数据解读", "item": "扣非净利润", "content": "", "evidence": "数据项11", "type": "分析"},
    {"seq": 9, "dim": "三、关键财务数据解读", "item": "毛利率", "content": "", "evidence": "数据项10,11", "type": "财报事实"},
    {"seq": 10, "dim": "三、关键财务数据解读", "item": "净利率", "content": "", "evidence": "数据项17", "type": "分析"},
    {"seq": 11, "dim": "三、关键财务数据解读", "item": "经营现金流", "content": "", "evidence": "N/A", "type": "数据缺口"},
    {"seq": 12, "dim": "三、关键财务数据解读", "item": "资产负债率", "content": "", "evidence": "数据项15", "type": "财报事实"},
    {"seq": 13, "dim": "三、关键财务数据解读", "item": "应收账款", "content": "", "evidence": "N/A", "type": "数据缺口"},
    {"seq": 14, "dim": "三、关键财务数据解读", "item": "存货", "content": "", "evidence": "N/A", "type": "数据缺口"},
    {"seq": 15, "dim": "三、关键财务数据解读", "item": "ROE", "content": "", "evidence": "数据项13", "type": "财报事实"},
    {"seq": 16, "dim": "三、关键财务数据解读", "item": "分红/回购/融资", "content": "", "evidence": "数据项2,21", "type": "财报事实"},
    
    # 四、识别隐藏风险 (5项)
    {"seq": 17, "dim": "四、识别隐藏风险", "item": "利润质量", "content": "", "evidence": "数据项11,17", "type": "分析"},
    {"seq": 18, "dim": "四、识别隐藏风险", "item": "现金流恶化", "content": "", "evidence": "N/A", "type": "推测"},
    {"seq": 19, "dim": "四、识别隐藏风险", "item": "资产风险", "content": "", "evidence": "数据项9", "type": "分析"},
    {"seq": 20, "dim": "四、识别隐藏风险", "item": "一次性收益", "content": "", "evidence": "数据项17", "type": "分析"},
    {"seq": 21, "dim": "四、识别隐藏风险", "item": "易忽略点", "content": "", "evidence": "数据项3,5", "type": "分析"},
    
    # 五、未来展望 (3项)
    {"seq": 22, "dim": "五、未来展望", "item": "增长逻辑", "content": "", "evidence": "数据项5", "type": "分析"},
    {"seq": 23, "dim": "五、未来展望", "item": "风险因素", "content": "", "evidence": "数据项5", "type": "推测"},
    {"seq": 24, "dim": "五、未来展望", "item": "跟踪指标", "content": "", "evidence": "数据项3,5", "type": "分析"},
    
    # 六、投资者视角结论 (3项)
    {"seq": 25, "dim": "六、投资者视角结论", "item": "长期投资者", "content": "", "evidence": "数据项2", "type": "分析"},
    {"seq": 26, "dim": "六、投资者视角结论", "item": "短期投资者", "content": "", "evidence": "数据项2", "type": "分析"},
    {"seq": 27, "dim": "六、投资者视角结论", "item": "机会 vs 风险", "content": "", "evidence": "数据项5,17", "type": "分析"},
]


# =====================================================================
# 主函数
# =====================================================================

def generate_analysis_report(
    company_name_cn: str,
    company_name_en: str,
    stock_code: str,
    fiscal_year: int,
    data_items: list,  # 数据表1的22项数据
    output_path: str,
    data_date: str = None
):
    """
    生成财报解读分析表（数据表2）
    
    参数：
    - company_name_cn: 公司中文名
    - company_name_en: 公司英文名
    - stock_code: 股票代码
    - fiscal_year: 财年
    - data_items: 数据表1的22项数据列表，格式: [[序号, 数据项, 具体值, 数据来源], ...]
    - output_path: 输出文件路径
    - data_date: 数据日期
    """
    # 构建数据索引（方便引用）
    data_dict = {}
    for item in data_items:
        if len(item) >= 3:
            seq = item[0]
            value = item[2]
            data_dict[f"数据项{seq}"] = value
    
    # 生成分析内容（基于模板 + 数据）
    analysis_list = []
    for template in ANALYSIS_TEMPLATE:
        seq = template["seq"]
        dim = template["dim"]
        item = template["item"]
        evidence = template["evidence"]
        type_ = template["type"]
        
        # 生成内容（基于数据项引用）
        content = generate_content(item, evidence, data_dict)
        
        analysis_list.append([seq, dim, item, content, evidence, type_])
    
    # 生成Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{company_name_cn} FY{fiscal_year} 财报解读'
    
    # 第1行: 标题
    ws.merge_cells('A1:F1')
    title_text = f'{company_name_cn}（{company_name_en}）({stock_code}) - FY{fiscal_year} 财报解读分析表'
    ws['A1'] = title_text
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = THIN_BORDER
    ws.row_dimensions[1].height = 36
    
    # 第2行: 元信息
    if data_date is None:
        data_date = datetime.now().strftime('%Y-%m-%d')
    ws.merge_cells('A2:F2')
    ws['A2'] = f'数据日期: {data_date} | 分析框架: 6维度27项 | 数据来源: 数据表1（按五级顺序查询，先查到先使用）'
    ws['A2'].font = DATE_FONT
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].border = THIN_BORDER
    
    # 第3行: 表头
    headers = ['序号', '维度', '分析项', '内容', '依据', '类型']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    # 数据行
    for row_idx, data in enumerate(analysis_list, 4):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal='left' if col_idx > 1 else 'center', 
                                        vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            # 交替行颜色
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
            else:
                cell.fill = DATA_FILL
    
    # 列宽设置
    ws.column_dimensions['A'].width = 5.0
    ws.column_dimensions['B'].width = 18.0
    ws.column_dimensions['C'].width = 15.0
    ws.column_dimensions['D'].width = 60.0
    ws.column_dimensions['E'].width = 15.0
    ws.column_dimensions['F'].width = 12.0
    
    # 行高设置
    for row in range(4, 4 + len(analysis_list)):
        ws.row_dimensions[row].height = 80
    
    # 冻结窗格
    ws.freeze_panes = 'A4'
    
    # 保存
    wb.save(output_path)
    print(f'[OK] 财报解读分析表已生成: {output_path}')
    return output_path


def generate_content(item_name: str, evidence: str, data_dict: dict) -> str:
    """根据分析项名称和数据字典生成可交付的基础分析内容。数据表1中的字段值已按五级顺序查询并在首次命中时写入。"""
    def get(key, default='数据表1未提供该项完整数据'):
        value = data_dict.get(key, default)
        if value is None or str(value).strip() in ['', '[待补充]', '待补充']:
            return default
        return str(value)

    main_business = get('数据项3')
    growth = get('数据项17')
    gross_margin = get('数据项11')
    roe = get('数据项13')
    debt = get('数据项15')
    capex = get('数据项9')
    future = get('数据项5')
    valuation = get('数据项2')
    buyback = get('数据项21')
    receivable_hint = get('数据项15')

    content_templates = {
        "整体判断": f"整体判断: 需结合增长、盈利能力、负债与现金回报综合看待。核心依据包括营收/利润变化({growth})、毛利率({gross_margin})、ROE({roe})和负债率({debt})。",
        "核心原因": f"核心原因: 1. 主营业务结构决定利润来源({main_business}); 2. 盈利能力由毛利率和ROE体现({gross_margin}; {roe}); 3. 财务安全边际由负债率、分红/回购和现金流相关信息体现({debt}; {buyback})。",
        "主营业务": f"公司主要依靠以下业务赚钱: {main_business}。投资者应重点观察主营业务收入占比、增速和毛利率变化。",
        "利润来源": f"利润来源主要来自高毛利或规模化业务。当前可见毛利率/利润能力线索为: {gross_margin}; ROE线索为: {roe}。",
        "经营变化": f"经营变化应重点从主营业务、收入增长和未来增长逻辑判断。当前数据表显示: 主营业务({main_business}); 增长情况({growth}); 未来逻辑({future})。",
        "营收": f"营收解读: {growth}。若收入保持增长且利润率稳定,通常说明主营业务需求和商业化能力较稳; 若增速放缓,需结合费用和现金流判断。",
        "归母净利润": f"归母净利润解读: {growth}。需要重点比较净利润增速与营收增速,若利润增速高于营收,通常说明经营杠杆或利润率改善。",
        "扣非净利润": f"扣非/核心利润解读: 数据表1未单列扣非净利润时,可先用毛利率、ROE和净利润变化辅助判断核心盈利质量。当前依据: 毛利率({gross_margin}); ROE({roe}); 增长({growth})。",
        "毛利率": f"毛利率解读: {gross_margin}。毛利率越稳定,通常说明产品/服务竞争力和成本控制越稳; 毛利率下降则需关注竞争、渠道分成、版权或原材料成本压力。",
        "净利率": f"净利率解读: 数据表1未单列净利率时,应结合营收、净利润和费用变化计算。当前可用依据: {growth}。",
        "经营现金流": f"经营现金流解读: 若数据表1未直接披露经营现金流,需回到现金流量表核验。可结合资本开支({capex})与利润增长({growth})判断现金转化质量。",
        "资产负债率": f"资产负债率解读: {debt}。负债率较低通常代表财务弹性较强; 负债率较高则需关注有息负债、合同负债和现金覆盖能力。",
        "应收账款": f"应收账款解读: 数据表1未单列应收账款时,需查资产负债表补充。可先结合公司负债率和业务模式判断回款风险。相关线索: {receivable_hint}。",
        "存货": f"存货解读: 数据表1未单列存货时,需查资产负债表补充。轻资产互联网/软件公司通常存货压力较低,制造业和内容版权型公司需重点关注跌价或减值风险。",
        "ROE": f"ROE解读: {roe}。ROE反映股东资本回报率,高ROE通常说明盈利能力或资产效率较强,但需结合负债率判断是否由高杠杆驱动。",
        "分红/回购/融资": f"股东回报解读: {buyback}。分红和回购可增强股东回报,但也要观察是否影响研发、资本开支和长期增长投入。",
        "利润质量": f"利润质量分析: 重点比较利润增长、毛利率、ROE与经营现金流是否匹配。当前依据: {growth}; {gross_margin}; {roe}。",
        "现金流恶化": f"现金流风险分析: 若利润增长但经营现金流未同步增长,需警惕回款、预付款或存货压力。当前需结合现金流量表与资本开支({capex})继续核验。",
        "资产风险": f"资产风险分析: 重点关注应收账款、存货、商誉/长期投资和合同负债。当前资产负债线索: {debt}; 资本开支/长期投入线索: {capex}。",
        "一次性收益": f"一次性收益分析: 需核对投资收益、公允价值变动、汇兑损益、资产减值等非经常性项目。当前数据表1未充分披露时,应在年报附注中继续核验。",
        "易忽略点": f"易忽略点: 投资者除收入和净利润外,还应关注业务结构({main_business})、未来增长逻辑({future})、资本开支({capex})和股东回报({buyback})。",
        "增长逻辑": f"增长逻辑: {future}。需要验证该增长逻辑是否能转化为收入增长、利润率稳定和现金流改善。",
        "风险因素": f"风险因素: 主要包括主营业务增速不及预期、毛利率下降、行业竞争加剧、资本开支回报不确定、估值偏高或股东回报下降。当前依据: {main_business}; {gross_margin}; {future}。",
        "跟踪指标": f"后续跟踪指标: 1. 主营业务收入增速; 2. 毛利率和ROE; 3. 经营现金流; 4. 资本开支回报; 5. 分红/回购执行; 6. 估值变化。当前依据: {growth}; {gross_margin}; {roe}; {buyback}。",
        "长期投资者": f"长期投资者视角: 若公司能维持主营业务竞争力、较高ROE、稳健负债率和清晰增长逻辑,长期价值更有支撑。当前依据: {main_business}; {roe}; {debt}; {future}。",
        "短期投资者": f"短期投资者视角: 更应关注业绩增速、市场预期差、分红回购和估值变化。当前估值/回报线索: {valuation}; {buyback}; 增长线索: {growth}。",
        "机会 vs 风险": f"机会与风险: 机会来自主营业务增长、盈利能力和股东回报; 风险来自增长放缓、利润率下行、现金流不足或估值过高。当前依据: {growth}; {gross_margin}; {roe}; {future}。",
    }

    return content_templates.get(item_name, '数据表1信息不足,需结合年报附注继续补充。')


def main():
    parser = argparse.ArgumentParser(
        description='财报解读分析表生成器 (数据表2)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用示例:
  python generate_analysis_report.py --company=福耀玻璃 --code=600660.SH --year=2025 --input=年报数据.xlsx --output=财报解读.xlsx

前提条件:
  1. 已生成数据表1（22项年报数据.xlsx）
  2. 数据表1包含完整的22项数据
        '''
    )
    parser.add_argument('--company', required=True, help='公司名称（中文）')
    parser.add_argument('--company-en', default='', help='公司英文名称')
    parser.add_argument('--code', required=True, help='股票代码')
    parser.add_argument('--year', type=int, required=True, help='财务年度')
    parser.add_argument('--input', required=True, help='数据表1路径（22项年报数据.xlsx）')
    parser.add_argument('--output', required=True, help='输出路径（财报解读.xlsx）')
    
    args = parser.parse_args()
    
    # 读取数据表1
    print(f'\n[Step 1] 读取数据表1: {args.input}')
    if not os.path.exists(args.input):
        print(f'[ERROR] 文件不存在: {args.input}')
        return
    
    # 使用openpyxl读取标准数据表1,跳过标题/元信息/表头行
    try:
        wb = openpyxl.load_workbook(args.input, data_only=True)
        ws = wb.active
        data_items = []
        for row in ws.iter_rows(min_row=4, max_row=25, values_only=True):
            if row and row[0] is not None:
                data_items.append(list(row[:4]))
        wb.close()
        print(f'  [OK] 读取到 {len(data_items)} 项数据')
    except Exception as e:
        print(f'[ERROR] 读取Excel失败: {e}')
        return
    
    # 生成分析表
    print(f'\n[Step 2] 生成财报解读分析表...')
    generate_analysis_report(
        company_name_cn=args.company,
        company_name_en=args.company_en or args.company,
        stock_code=args.code,
        fiscal_year=args.year,
        data_items=data_items,
        output_path=args.output
    )
    
    print(f'\n[完成] 输出文件: {args.output}')


if __name__ == '__main__':
    main()
