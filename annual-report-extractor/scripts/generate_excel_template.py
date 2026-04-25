# -*- coding: utf-8 -*-
"""
年报22项数据Excel生成脚本
annual-report-extractor 技能御用模板生成器

严格按照 annual-report-extractor/SKILL.md 的样式规范：
- Sheet名: {公司中文名} FY{年份} 22项数据
- 标题绿: 006B5A | 表头浅绿: 00E8F5F0 | 数据白: FFFFFF
- 列宽: A:5.0 / B:38.0 / C:65.0 / D:45.0
- 行高: 数据行100像素
- 冻结: A4
- 禁止emoji，用纯文字标注数据来源和可靠性

使用方法:
python generate_excel_template.py

Author: QClaw
Date: 2026-04-23
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import os
from datetime import datetime


# =====================================================================
# 样式常量（固定，不得更改）
# =====================================================================
TITLE_FILL = PatternFill('solid', fgColor='006B5A')   # 绿城绿-标题
HEADER_FILL = PatternFill('solid', fgColor='00E8F5F0')  # 浅绿-表头
DATA_FILL = PatternFill('solid', fgColor='FFFFFF')    # 白色-数据行
ALT_FILL = PatternFill('solid', fgColor='F5FAFA')     # 淡绿-交替行

TITLE_FONT = Font(name='Arial', size=12, bold=True, color='FFFFFF')
HEADER_FONT = Font(name='Arial', size=10, bold=False, color='000000')
DATA_FONT = Font(name='Arial', size=10, bold=False, color='000000')
DATE_FONT = Font(name='Arial', size=9, color='666666')

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


# =====================================================================
# 主函数
# =====================================================================

def generate_annual_report_excel(
    company_name_cn: str,
    company_name_en: str,
    stock_code: str,
    fiscal_year: int,
    currency_unit: str,
    data_list: list,
    output_path: str,
    data_date: str = None
):
    """
    生成年报22项数据Excel文件
    
    参数：
    - company_name_cn: 公司中文名（如：福耀玻璃）
    - company_name_en: 公司英文名（如：Fuyao Glass Industry）
    - stock_code: 股票代码（如：600660.SH）
    - fiscal_year: 财年（如：2025）
    - currency_unit: 货币单位（如：亿元人民币）
    - data_list: 22项数据列表
      格式: [[序号, 数据项, 具体值, 数据来源], ...]
      序号1-22，对应SKILL.md中的22项清单
    - output_path: 输出文件路径（推荐: {公司名}_FY{年份}_22项数据.xlsx）
    - data_date: 数据截止日期，默认为当天
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Sheet名：22项（统一规范）
    ws.title = f'{company_name_cn} FY{fiscal_year} 22项数据'
    
    # ====== 第1行：标题 ======
    ws.merge_cells('A1:D1')
    title_text = (
        f'{company_name_cn}（{company_name_en}）({stock_code})'
        f' - FY{fiscal_year} 年报 22项核心数据汇总'
    )
    ws['A1'] = title_text
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = THIN_BORDER
    ws.row_dimensions[1].height = 36
    
    # ====== 第2行：数据日期/元信息 ======
    if data_date is None:
        data_date = datetime.now().strftime('%Y-%m-%d')
    ws.merge_cells('A2:D2')
    meta_text = (
        f'数据日期：{data_date} | '
        f'货币单位：{currency_unit}（除注明外） | '
        f'财年：FY{fiscal_year}（截至{fiscal_year}年12月31日） | '
        f'数据来源见D列'
    )
    ws['A2'] = meta_text
    ws['A2'].font = DATE_FONT
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 26
    
    # ====== 第3行：表头 ======
    headers = ['序号', '数据项', '具体值', '数据来源/说明']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 26
    
    # ====== 数据行（第4行起）======
    for i, row_data in enumerate(data_list):
        row = i + 4
        fill = DATA_FILL if i % 2 == 0 else ALT_FILL
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=False)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 100  # 行高100像素
    
    # ====== 列宽（固定，不得更改）======
    ws.column_dimensions['A'].width = 5.0    # 序号
    ws.column_dimensions['B'].width = 38.0  # 数据项
    ws.column_dimensions['C'].width = 65.0  # 具体值
    ws.column_dimensions['D'].width = 45.0  # 数据来源
    
    # ====== 冻结窗格 ======
    ws.freeze_panes = 'A4'
    
    # ====== 保存 ======
    wb.save(output_path)
    print(f'[OK] Excel已生成: {output_path}')
    print(f'     Sheet名: {ws.title}')
    print(f'     数据行: {len(data_list)} 项')
    return output_path


# =====================================================================
# 22项标准数据模板（空模板）
# =====================================================================

def create_empty_template() -> list:
    """返回22项空数据模板，供填充使用"""
    return [
        [1,  '公司名字',              '[待填充]',                                              '[待填充]'],
        [2,  '公司市值',              '[待填充]',                                              '[待填充]'],
        [3,  '公司主营业务',          '[待填充]',                                              '[待填充]'],
        [4,  '市场份额',              '[待填充]',                                              '[待填充]'],
        [5,  '未来增长率',            '[待填充]',                                              '[待填充]'],
        [6,  '上游供应商',            '[待填充]',                                              '[待填充]'],
        [7,  '下游客户',              '[待填充]',                                              '[待填充]'],
        [8,  '原材料成本',            '[待填充]',                                              '[待填充]'],
        [9,  '资本开支',              '[待填充]',                                              '[待填充]'],
        [10, '行业平均毛利率',        '[待填充]',                                              '[待填充]'],
        [11, '公司毛利率',            '[待填充]',                                              '[待填充]'],
        [12, '行业平均ROE',           '[待填充]',                                              '[待填充]'],
        [13, '公司ROE',               '[待填充]',                                              '[待填充]'],
        [14, '行业平均负债率',        '[待填充]',                                              '[待填充]'],
        [15, '公司负债率',            '[待填充]',                                              '[待填充]'],
        [16, '合同负债',              '[待填充]',                                              '[待填充]'],
        [17, '营收增长率',            '[待填充]',                                              '[待填充]'],
        [18, 'PE百分位',              '[待填充]',                                              '[待填充]'],
        [19, 'PB百分位',              '[待填充]',                                              '[待填充]'],
        [20, '美股同类公司',          '[待填充]',                                              '[待填充]'],
        [21, '股票增减持',            '[待填充]',                                              '[待填充]'],
        [22, '高管增减持',            '[待填充]',                                              '[待填充]'],
    ]


# =====================================================================
# 示例用法
# =====================================================================

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='生成标准年报22项数据Excel模板')
    parser.add_argument('--output', help='模板输出路径。不提供时只打印用法,不生成文件。')
    parser.add_argument('--company', default='示例公司', help='公司中文名')
    parser.add_argument('--company-en', default='Example Corp', help='公司英文名')
    parser.add_argument('--code', default='000000.SZ', help='股票代码')
    parser.add_argument('--year', type=int, default=2025, help='财年')
    args = parser.parse_args()

    print('[annual-report-extractor] Excel生成脚本')
    print('使用示例:')
    print('  python generate_excel_template.py --output template.xlsx')
    print('  from generate_excel_template import generate_annual_report_excel, create_empty_template')
    print()

    if not args.output:
        print('[INFO] 未提供 --output,不会生成模板文件。')
        raise SystemExit(0)

    output_dir = os.path.dirname(os.path.abspath(args.output))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    empty = create_empty_template()
    generate_annual_report_excel(
        company_name_cn=args.company,
        company_name_en=args.company_en,
        stock_code=args.code,
        fiscal_year=args.year,
        currency_unit='亿元人民币',
        data_list=empty,
        output_path=args.output,
        data_date=datetime.now().strftime('%Y-%m-%d')
    )
    print(f'\n[OK] 模板文件已生成: {args.output}')
