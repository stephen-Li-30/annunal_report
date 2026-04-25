# -*- coding: utf-8 -*-
import re

path = r'C:\Users\lishaoming\.qclaw\workspace\skills\annual-report-extractor\SKILL.md'
with open(path, 'r', encoding='utf-8') as f:
    content = f.read()

# ======= 替换1：脚本速查表（原3列→4列）=======
old_table = """### 脚本速查

| 脚本 | 用途 | 使用方式 |
|------|------|---------|
| `check_env.py` | 环境检查 | `python check_env.py` |
| `extract_data.py` | 一键提取（AKShare→6轮搜索→Excel→验证） | `python extract_data.py --company=公司名 --code=代码 --year=年份` |
| `akshare_universal_integration.py` | AKShare全市场数据提取 | `python akshare_universal_integration.py` |
| `generate_excel_template.py` | 生成空模板 | `python generate_excel_template.py` |
| `data_validator.py` | 数据验证 | `python data_validator.py <excel文件>` |
| `data_validator_enhanced.py` | 五维度数据验证 | `python data_validator_enhanced.py` |"""

new_table = """### 脚本速查

| 脚本 | 用途 | 主要依赖 | 使用方式 |
|------|------|---------|----------|
| `check_env.py` | 环境检查 | - | `python check_env.py` |
| `extract_data.py` | 一键提取（AKShare→6轮搜索→Excel→验证） | akshare, pandas | `python extract_data.py --company=公司名 --code=代码 --year=年份` |
| `akshare_universal_integration.py` | AKShare全市场数据提取 | akshare, pandas | `python akshare_universal_integration.py` |
| `generate_excel_template.py` | 生成空模板 | openpyxl, pandas | `python generate_excel_template.py` |
| `data_validator.py` | 数据验证 | openpyxl, pandas | `python data_validator.py <excel文件>` |
| `data_validator_enhanced.py` | 五维度数据验证 | pandas | `python data_validator_enhanced.py` |
| `read_excel.py` | **Pandas读取Excel** | pandas, openpyxl | `python read_excel.py <excel文件>` |"""

if old_table in content:
    content = content.replace(old_table, new_table)
    print('[OK] 脚本速查表已更新')
else:
    print('[WARN] 脚本速查表未找到精确匹配，跳过')

# ======= 替换2：在"## 标准工作流程"前插入Pandas章节=======
pandas_section = """## Pandas 数据处理能力

本技能全面集成 pandas 用于 Excel 读写、数据分析和报告生成。

### 安装确认

```bash
pip show pandas   # 查看已安装版本
pip install pandas openpyxl  # 如未安装，执行此命令
```

### 读取 Excel 二进制文件（.xlsx）

```python
import pandas as pd

# 读取整个Sheet（默认第一个Sheet）
df = pd.read_excel('年报数据.xlsx', sheet_name='Sheet1')

# 读取指定Sheet
df = pd.read_excel('年报数据.xlsx', sheet_name='数据验证')

# 读取所有Sheet（返回字典）
sheets = pd.read_excel('年报数据.xlsx', sheet_name=None)
for name, data in sheets.items():
    print(f"Sheet: {name}, 行数: {len(data)}")

# 跳过标题行，指定header行
df = pd.read_excel('年报数据.xlsx', header=2)

# 只读取指定列
df = pd.read_excel('年报数据.xlsx', usecols=['公司名称', '营收', '净利润'])
```

### 数据分析常用操作

```python
import pandas as pd

df = pd.read_excel('年报数据.xlsx')

# 基础统计
print(df.describe())           # 数值列统计摘要
print(df.dtypes)               # 列类型
print(df.isnull().sum())       # 缺失值统计

# 筛选
high_roe = df[df['ROE'] > 15]           # ROE > 15%
revenue_growth = df[df['营收增长率'] > 0]  # 正增长公司

# 排序
df_sorted = df.sort_values('ROE', ascending=False)

# 分组聚合
industry_avg = df.groupby('行业')['毛利率'].mean()

# 新增计算列
df['净利率'] = df['净利润'] / df['营收'] * 100

# 导出结果
df.to_excel('分析结果.xlsx', index=False)
```

### 与 openpyxl 联合使用（样式+数据）

```python
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Step 1: 用 pandas 读取数据
df = pd.read_excel('年报数据.xlsx')

# Step 2: 数据分析
analysis = df[['公司名称', '营收', '净利润', 'ROE', '负债率']].copy()
analysis['净利率'] = analysis['净利润'] / analysis['营收'] * 100

# Step 3: 用 openpyxl 写入并保留样式
with pd.ExcelWriter('输出.xlsx', engine='openpyxl') as writer:
    analysis.to_excel(writer, sheet_name='分析结果', index=False)

    # 打开工作簿添加样式
    wb = writer.book
    ws = wb['分析结果']

    # 标题行绿色背景
    green_fill = PatternFill(start_color='006B5A', end_color='006B5A', fill_type='solid')
    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = Font(color='FFFFFF', bold=True)
```

### 快速验证 Excel 数据结构

```python
import pandas as pd

def validate_excel_structure(path):
    \"\"\"快速验证Excel是否符合22项数据格式\"\"\"
    df = pd.read_excel(path)

    print(f"Sheet: {path}")
    print(f"行数: {len(df)}, 列数: {len(df.columns)}")
    print(f"列名: {list(df.columns)}")
    print(f"\\n前5行预览:")
    print(df.head())

    # 检查空值
    nulls = df.isnull().sum()
    if nulls.any():
        print(f"\\n存在空值的列:")
        print(nulls[nulls > 0])

# 使用
validate_excel_structure('年报数据.xlsx')
```

### read_excel.py 脚本（内置·开箱即用）

技能包内置 `scripts/read_excel.py`，直接读取并展示 Excel 内容：

```bash
# 读取并打印 Excel 全部内容
python scripts/read_excel.py 年报数据.xlsx

# 读取指定 Sheet
python scripts/read_excel.py 年报数据.xlsx --sheet 数据验证

# 只显示前10行
python scripts/read_excel.py 年报数据.xlsx --n 10

# 检查数据完整性（检测空值）
python scripts/read_excel.py 年报数据.xlsx --check-null
```

---

"""

marker = "## 标准工作流程"
if marker in content:
    content = content.replace(marker, pandas_section + marker)
    print('[OK] Pandas章节已插入到"## 标准工作流程"之前')
else:
    print('[WARN] 未找到"## 标准工作流程"标记')

with open(path, 'w', encoding='utf-8') as f:
    f.write(content)

print('文件已保存')
